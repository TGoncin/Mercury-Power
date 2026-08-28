#WriteXlsx.py
import shutil
from pathlib import Path
from sqlite3 import Row
from Resources.resources import resource_path
from openpyxl import load_workbook
import xlwings as xw
import os
from datetime import datetime
from enum import Enum
from Definitions.enums import eConnectionType,eArrayType,eLogical,eBIL,eTransformerFamily,eWindingSelection,eTapExtremes,eSequence,eChangerCapacity
from Definitions.enums import eMaterial,eLaminationType,eTypeConductor,eCoating,eDesignSpec,eCoolingType,eSingleCoolingType,eTemperatureRise,eModelChanger
from Definitions.enums import eStandardTemperatureRise,eStandardLightningImpulse,eStandardRIV,eStandardSoundLevel,eStandardPartialDischarge,eLeads,eBars,eLaminationTypeNum
from Definitions.functions import address,ADDRESSZ,is_folder_open
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image

def save_excel(device):
    try:
        """
        Escribe todos los valores del 'device' en el archivo maestro
        y lo guarda en la carpeta de diseño según sAddress
        """

        iNoPhases = device.fields["iNoPhases"].get()

        #Archivo maestro
        if iNoPhases==1:
            master_file = resource_path("Resources/1 Ø L-H-L Electrical Design Master.xlsx")
        else:   
            master_file = resource_path("Resources/3 Ø L-H Electrical Design Master.xlsx")

        documents_folder = Path.home() / "Documents"
        so_number = str(device.fields["so_number"].get())
        dest_folder = documents_folder / so_number
        dest_folder.mkdir(parents=True, exist_ok=True)

        
        output_file = dest_folder / f"{so_number}.xlsx"

        # 4️⃣ Copy exact file
        shutil.copy2(master_file, output_file)


        wb = load_workbook(output_file)
        
        for ws in wb.worksheets:
            ws.protection.sheet = False
            ws.protection.enable = False

        # Desproteger estructura del workbook
        #wb.security.workbookPassword = None
        #wb.security.lockStructure = False

        sheet = wb.active

        if iNoPhases==1:
           single_phase(sheet,device)
        else:
            three_phases(sheet,device)

        # ---------------------------
        # PRICES
        # ---------------------------

        if "Prices" in wb.sheetnames:
            del wb["Prices"]

        wsP = wb.create_sheet("Prices")

        if not is_folder_open(dest_folder):
            os.startfile(dest_folder)

        current_row = 1

        def write_table(W,current_row,headers, data, title):
    
            # Title
            W.cell(row=current_row, column=1, value=title)
            current_row += 1
    
            # Headers
            for col, header in enumerate(headers, start=1):
                W.cell(row=current_row, column=col, value=header)
            current_row += 1
    
            # Data
            for row_data in data:
                for col, value in enumerate(row_data, start=1):
                    W.cell(row=current_row, column=col, value=value)
                current_row += 1
    
            current_row += 2 
            return current_row

        dCostLam = device.costs.fields["dLamination"].get()
        wsP.cell(row=current_row, column=1, value = "CoreGrade")
        wsP.cell(row=current_row, column=2, value = "Price")
        current_row +=1 
        wsP.cell(row=current_row, column=1, value = "Am")
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.Am.value])
        current_row +=1 
        wsP.cell(row=current_row, column=1, value = "Am HB1")
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.AmHB1.value])
        current_row +=1 
        wsP.cell(row=current_row, column=1, value = "MOH")
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.M0H.value])
        current_row +=1 
        wsP.cell(row=current_row, column=1, value = "M2") 
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.M2.value])
        current_row +=1    
        wsP.cell(row=current_row, column=1, value = "M3")
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.M3.value])
        current_row +=1 
        wsP.cell(row=current_row, column=1, value = "M4")
        wsP.cell(row=current_row, column=2, value = dCostLam[eLaminationTypeNum.M4.value])
        current_row +=3

        # current_row = write_table(wsP,current_row,
        #     device.tbConductor.tb_headers_magnet,
        #     device.tbConductor.tb_data_COPPER_magnet,
        #     "COPPER MAGNET"
        # )

        # current_row= write_table(wsP,current_row,
        #     device.tbConductor.tb_headers_strip,
        #     device.tbConductor.tb_data_ALUMINUM_strip,
        #     "ALUMINUM STRIP"
        # )

        # current_row = write_table(wsP,current_row,
        #     device.tbConductor.tb_headers_strip,
        #     device.tbConductor.tb_data_COPPER_strip,
        #     "COPPER STRIP"
        # )

        # current_row = write_table(wsP,current_row,
        #     device.tbConductor.tb_headers_magnet,
        #     device.tbConductor.tb_data_ALUMINUM_magnet,
        #     "ALUMINUM MAGNET"
        # )
    
        current_row = write_table(wsP,current_row,
            device.SAP_STD_INS_CORE_LEADS.tb_headers,
            device.SAP_STD_INS_CORE_LEADS.tb_data,
            "SAP STD INS CORE LEADS"
        )

        iNoPhases = device.fields["iNoPhases"].get()
        tLamination = device.core.fields["tLaminationType"].get()
        bAmorphous = True if tLamination == eLaminationType.Am or tLamination == eLaminationType.AmHB1 else False

        if iNoPhases==1:
            if bAmorphous:
                img_path_core = resource_path("Resources/single_phase_core_Am.png")
            else:
                img_path_core = resource_path("Resources/single_phase_core_M0H.png")
        else:
            if bAmorphous:
                img_path_core = resource_path("Resources/three_phases_core_Am.png")
            else:
                img_path_core = resource_path("Resources/three_phases_core_M0H.png")
        img_pil_core = PILImage.open(img_path_core)
        dpi_core = 96
        if iNoPhases==1:
            width_inch_core = 4
            height_inch_core = 2.2
        else:
            width_inch_core = 6
            height_inch_core = 2.2
        width_px_core = int(width_inch_core * dpi_core)
        height_px_core = int(height_inch_core * dpi_core)

        img_pil_core = img_pil_core.resize((width_px_core, height_px_core), PILImage.Resampling.LANCZOS)
        img_bytes_core = BytesIO()
        img_pil_core.save(img_bytes_core, format="PNG")
        img_bytes_core.seek(0)

        img_core = Image(img_bytes_core)
        if iNoPhases==3:
            sheet.add_image(img_core, f"D224")  
        else:
            sheet.add_image(img_core, f"D211")  

        def pil_to_xl_image(pil_img):
            buffer = BytesIO()
            pil_img.save(buffer, format="PNG")
            return Image(BytesIO(buffer.getvalue()))

        def remove_white_background(img_path, threshold=240):
            img = PILImage.open(img_path).convert("RGBA")
            datas = img.getdata()

            new_data = []
            for r, g, b, a in datas:
                if r > threshold and g > threshold and b > threshold:
                    new_data.append((255, 255, 255, 0))
                else:
                    new_data.append((r, g, b, 255))

            img.putdata(new_data)
            return img

        #logo image
        img_path_logo = resource_path("Resources/logo_reports_ori.png")
        img_pil_logo = remove_white_background(img_path_logo)
        #img_pil_logo = PILImage.open(img_path_logo)
        dpi_logo = 96
        if iNoPhases==1:
            width_inch_logo = 1.98
            height_inch_logo = 0.356
        else:
            width_inch_logo = 1.98
            height_inch_logo = 0.356
        width_px_logo = int(width_inch_logo * dpi_logo)
        height_px_logo = int(height_inch_logo * dpi_logo)

        img_pil_logo = img_pil_logo.resize((width_px_logo, height_px_logo), PILImage.Resampling.LANCZOS)

        if iNoPhases == 3:
            logo_cells = ["J140", "J212", "J277", "J341", "J404", "J467", "J532", "J596", "J667", "J733"]
            for cell in logo_cells:
                sheet.add_image(pil_to_xl_image(img_pil_logo), cell)
        else:
            logo_cells = ["J135", "J202", "J265", "J328", "J391", "J454", "J517", "J580", "J643", "J710"]
            for cell in logo_cells:
                sheet.add_image(pil_to_xl_image(img_pil_logo), cell) 

        if iNoPhases == 3:
            #INSULATION B380
            img_path_ID_B380 = resource_path("Resources/ID_B380.png")
            img_pil_ID_B380 = remove_white_background(img_path_ID_B380)
            #img_pil_ID_B380 = PILImage.open(img_path_ID_B380)
            dpi_ID_B380 = 96
            width_inch_ID_B380 = 4.664
            height_inch_ID_B380 = 0.72
            width_px_ID_B380 = int(width_inch_ID_B380 * dpi_ID_B380)
            height_px_ID_B380 = int(height_inch_ID_B380 * dpi_ID_B380)
            img_pil_ID_B380 = img_pil_ID_B380.resize((width_px_ID_B380, height_px_ID_B380), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_B380), "B380")

            #INSULATION D391
            img_path_ID_D391 = resource_path("Resources/ID_D391.png")
            img_pil_ID_D391 = remove_white_background(img_path_ID_D391)
            #img_pil_ID_D391 = PILImage.open(img_path_ID_D391)
            dpi_ID_D391 = 96
            width_inch_ID_D391 = 2.698
            height_inch_ID_D391 = 0.525
            width_px_ID_D391 = int(width_inch_ID_D391 * dpi_ID_D391)
            height_px_ID_D391 = int(height_inch_ID_D391 * dpi_ID_D391)
            img_pil_ID_D391 = img_pil_ID_D391.resize((width_px_ID_D391, height_px_ID_D391), PILImage.Resampling.LANCZOS)

            sheet.add_image(pil_to_xl_image(img_pil_ID_D391), "D391")

            #INSULATION H376
            img_path_ID_H376 = resource_path("Resources/ID_H376.png")
            img_pil_ID_H376 = remove_white_background(img_path_ID_H376)
            #img_pil_ID_H376 = PILImage.open(img_path_ID_H376)
            dpi_ID_H376 = 96
            width_inch_ID_H376 = 4.047
            height_inch_ID_H376 = 0.787
            width_px_ID_H376 = int(width_inch_ID_H376 * dpi_ID_H376)
            height_px_ID_H376 = int(height_inch_ID_H376 * dpi_ID_H376)
            img_pil_ID_H376 = img_pil_ID_H376.resize((width_px_ID_H376, height_px_ID_H376), PILImage.Resampling.LANCZOS)

            sheet.add_image(pil_to_xl_image(img_pil_ID_H376), "H376")


            #FORM SHEET
            img_path_FS_C692 = resource_path("Resources/FS_C692.png")
            img_pil_FS_C692 = remove_white_background(img_path_FS_C692)
            #img_pil_FS_C692 = PILImage.open(img_path_FS_C692)
            sheet.add_image(pil_to_xl_image(img_pil_FS_C692), "B691")

        else:
            #INSULATION C358
            img_path_ID_C358 = resource_path("Resources/ID_C358SP.png")
            img_pil_ID_C358 = remove_white_background(img_path_ID_C358)
            img_pil_ID_C358 = PILImage.open(img_path_ID_C358)
            dpi_ID_C358 = 96
            width_inch_ID_C358 = 5.41
            height_inch_ID_C358 = 0.3779
            width_px_ID_C358 = int(width_inch_ID_C358 * dpi_ID_C358)
            height_px_ID_C358 = int(height_inch_ID_C358 * dpi_ID_C358)
            img_pil_ID_C358 = img_pil_ID_C358.resize((width_px_ID_C358, height_px_ID_C358), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_C358), "C359")

            #INSULATION C367
            img_path_ID_C367 = resource_path("Resources/ID_C367SP.png")
            img_pil_ID_C367 = remove_white_background(img_path_ID_C367)
            img_pil_ID_C367 = PILImage.open(img_path_ID_C367)
            dpi_ID_C367 = 96
            width_inch_ID_C367 = 3.0826
            height_inch_ID_C367 = 0.886
            width_px_ID_C367 = int(width_inch_ID_C367 * dpi_ID_C367)
            height_px_ID_C367 = int(height_inch_ID_C367 * dpi_ID_C367)
            img_pil_ID_C367 = img_pil_ID_C367.resize((width_px_ID_C367, height_px_ID_C367), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_C367), "C367")

            #INSULATION C378
            img_path_ID_C378 = resource_path("Resources/ID_C378SP.png")
            img_pil_ID_C378 = remove_white_background(img_path_ID_C378)
            img_pil_ID_C378 = PILImage.open(img_path_ID_C378)
            dpi_ID_C378 = 96
            width_inch_ID_C378 = 3.476
            height_inch_ID_C378 = 0.67
            width_px_ID_C378 = int(width_inch_ID_C378 * dpi_ID_C378)
            height_px_ID_C378 = int(height_inch_ID_C378 * dpi_ID_C378)
            img_pil_ID_C378 = img_pil_ID_C378.resize((width_px_ID_C378, height_px_ID_C378), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_C378), "C378")

            #INSULATION E349
            img_path_ID_E349 = resource_path("Resources/ID_E349SP.png")
            img_pil_ID_E349 = remove_white_background(img_path_ID_E349)
            img_pil_ID_E349 = PILImage.open(img_path_ID_E349)
            dpi_ID_E349 = 96
            width_inch_ID_E349 = 1.9055
            height_inch_ID_E349 = 1.8
            width_px_ID_E349 = int(width_inch_ID_E349 * dpi_ID_E349)
            height_px_ID_E349 = int(height_inch_ID_E349 * dpi_ID_E349)
            img_pil_ID_E349 = img_pil_ID_E349.resize((width_px_ID_E349, height_px_ID_E349), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_E349), "E348")

            #INSULATION H363
            img_path_ID_H363 = resource_path("Resources/ID_H363SP.png")
            img_pil_ID_H363 = remove_white_background(img_path_ID_H363)
            img_pil_ID_H363 = PILImage.open(img_path_ID_H363)
            dpi_ID_H363 = 96
            width_inch_ID_H363 = 3.5433
            height_inch_ID_H363 = 0.9212
            width_px_ID_H363 = int(width_inch_ID_H363 * dpi_ID_H363)
            height_px_ID_H363 = int(height_inch_ID_H363 * dpi_ID_H363)
            img_pil_ID_H363 = img_pil_ID_H363.resize((width_px_ID_H363, height_px_ID_H363), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_ID_H363), "H363")

            #FORM SHEET
            img_path_FS_B665 = resource_path("Resources/FS_B665SP.png")
            img_pil_FS_B665 = remove_white_background(img_path_FS_B665)
            #img_pil_FS_B665 = PILImage.open(img_path_FS_B665)
            sheet.add_image(pil_to_xl_image(img_pil_FS_B665), "B665")

            img_path_FS_B688 = resource_path("Resources/FS_B688SP.png")
            img_pil_FS_B688 = remove_white_background(img_path_FS_B688)
            dpi_ID_FS_B688 = 96
            width_inch_FS_B688 = 9.03
            height_inch_FS_B688 = 3.173
            width_px_FS_B688 = int(width_inch_FS_B688 * dpi_ID_FS_B688)
            height_px_FS_B688 = int(height_inch_FS_B688 * dpi_ID_FS_B688)
            img_pil_FS_B688 = img_pil_FS_B688.resize((width_px_FS_B688, height_px_FS_B688), PILImage.Resampling.LANCZOS)
            sheet.add_image(pil_to_xl_image(img_pil_FS_B688), "B688")

        # if "SAP" in wb.sheetnames:
        #     del wb["SAP"]

        # wsS = wb.create_sheet("SAP")
        # current_row = 1
        # current_row = write_table(wsS,current_row,
        #     device.SAP_STD_CONDUCTORS.tb_headers,
        #     device.SAP_STD_CONDUCTORS.tb_data,
        #     "SAP STD CONDUCTORS"
        # )

        # ---------------------------
        # Carpeta y guardado
        # ---------------------------
        wb.save(output_file)

        if not is_folder_open(dest_folder):
            os.startfile(dest_folder)

        return True
    except Exception as ex: 
        return False

def single_phase(sheet,device):
    # ---------------------------
    # GENERAL
    # ---------------------------
    sheet["P81"].value = device.fields["so_number"].get()
    sheet["P82"].value = device.fields["customer_name"].get()
    sheet["P83"].value = device.fields["customer_code"].get()
    sheet["P84"].value = device.fields["quantity"].get()
    sheet["P85"].value = device.guarantee.fields["dEvaluation_nl"].get()
    sheet["P86"].value = device.guarantee.fields["dEvaluation_ll"].get()
    sheet["P87"].value = device.fields["issued_by"].get()
    date_str = device.fields["original_issue_date"].get()  # '11/18/2025'

    date_obj = datetime.strptime(date_str, "%m/%d/%Y")

    sheet["P88"].value = date_obj
    #sheet["P88"].api.NumberFormat = "dd-mmm-yy"  

    sheet["P89"].value = device.guarantee.fields["sCustomerSpec"].get()
    sheet["P90"].value = device.guarantee.fields["tDesignSpec"].get().value
    sheet["P92"].value = "YES" if device.guarantee.fields["bNLCorrected_to_85C"].get()==eLogical.TRUE else "NO"
    sheet["P93"].value = device.guarantee.fields["dNL_pu"].get()
    sheet["P96"].value = device.fields["nominal_kva"].get()
    sheet["P97"].value = device.cooling.fields["tCoolingTypes"].get().value
    sheet["P98"].value = device.fields["frequency"].get()
    sheet["P99"].value = int(device.cooling.fields["tTemperatureRises"].get().name.split("_")[-1])  # Temp_65 → 65
    sheet["P100"].value = device.fields["transformer_family"].get().name
    sheet["P101"].value = device.fields["winding_description"].get()

    # ---------------------------
    # TAPS
    # ---------------------------
    pWindH = device.getHVWind()
    if pWindH.have_taps()==eLogical.TRUE:
        sheet["P104"].value = "Yes"
        changer = pWindH.changer
        sheet["P105"].value = float(changer.fields["dPerUp"].get()/changer.taps_sup())  # Porcentaje paso
        sheet["P106"].value = pWindH.fields["iNomPos"].get()  # nominal
        sheet["P107"].value = changer.get_NumSteps()+1   # Número de posiciones
        sheet["P109"].value = changer.fields["tCapacity"].get().name
        sheet["P110"].value = changer.fields["sInstructions"].get()
    else:
        sheet["P104"].value = "No"

    # ---------------------------
    # HV WINDING
    # ---------------------------
    sheet["P113"].value = pWindH.fields["dVoltage"].get()
    sheet["P114"].value = 1.73 if pWindH.fields["type_connection"].get().name == "DELTA" else 1
    sheet["P115"].value = int(pWindH.fields["iBIL"].get().name.split("_")[-1])
    sheet["P116"].value = "Yes" if pWindH.fields["ground_Y"].get() else "No"

    # ---------------------------
    # LV WINDING
    # ---------------------------
    pWindL = device.getLVWind()
    sheet["P118"].value = pWindL.fields["dVoltage"].get()
    sheet["P119"].value = 1.73 if pWindL.fields["type_connection"].get().name == "DELTA" else 1
    sheet["P120"].value = int(pWindL.fields["iBIL"].get().name.split("_")[-1])
    sheet["P121"].value = pWindL.fields["iNumBushings"].get()
    sheet["P123"].value = pWindL.fields["iQuantityLeads"].get()
    sheet["P124"].value = pWindL.fields["dThicknessLeads"].get()
    sheet["P125"].value = pWindL.fields["dWidthLeads"].get()

    sheet["P126"].value = pWindL.fields["tMaterialLeads"].get().name
    sheet["P127"].value = device.fields["iQuantityBuswork"].get()
    sheet["P128"].value = device.fields["dThicknessBuswork"].get()
    sheet["P129"].value = device.fields["dWidthBuswork"].get()
    sheet["P130"].value = device.fields["tMaterialBuswork"].get().name

    # ---------------------------
    # CORE
    # ---------------------------
    core = device.core
    sheet["P136"].value = core.fields["tLaminationType"].get().value
    sheet["P137"].value = core.fields["dTotalStack"].get()
    sheet["P138"].value = core.fields["dSteelWidth"].get()
    sheet["P139"].value = core.fields["dPerOvervoltageMaxFlux"].get()
    sheet["P140"].value = core.fields["dTestInductionLevel"].get()
    sheet["P141"].value = core.fields["dMaxWattsPound"].get()
    sheet["P142"].value = core.fields["dLossFact"].get()
    # ---------------------------
    # LV Winding
    # ---------------------------
    pWindL = device.getLVWind()
    pCoilL = pWindL.coils[0]
    pCondL = pCoilL.conductor[0]
    iType = pCondL.fields["iType"].get()

    sheet["P145"].value = pWindL.fields["dNomTurnsUI"].get()
    sheet["P146"].value = pCoilL.fields["iLayers"].get()
    sheet["P147"].value = 1 if iType==eTypeConductor.STRIP else 2
        
    sheet["P148"].value = pCondL.fields["iNoConductor"].get()
    sheet["P149"].value = pCondL.fields["dWidth"].get()
    sheet["P150"].value = pCondL.fields["iNoAxial"].get()
    sheet["P151"].value = pCondL.fields["iNoRadial"].get()
    
    sheet["P155"].value = pWindL.fields["dLeadBuild"].get()
    sheet["P156"].value = pWindL.fields["iNoDuctInner"].get()
    sheet["P157"].value = pWindL.fields["iNoDuctOuter"].get()
    sheet["P158"].value = pWindL.fields["dDuctSize"].get()
    sheet["P159"].value = pWindL.fields["tCoating"].get().value
    # ---------------------------
    # HV Winding
    # ---------------------------
    pWindH = device.getHVWind()
    pCoilH = pWindH.coils[0]
    pCondH = pCoilH.conductor[0]
    sheet["P162"].value = pCoilH.fields["bConstantStayback"].get().value
    sheet["P163"].value = pCondH.fields["iNoConductor"].get()
    sheet["P164"].value = pCondH.fields["iNoAxial"].get()
    sheet["P165"].value = pCondH.fields["iNoRadial"].get()
    sheet["P170"].value = pWindH.fields["dLeadBuild"].get()
    sheet["P171"].value = pWindH.fields["iNoDuctInner"].get()   
    # ---------------------------
    # REVISION DATA
    # ---------------------------
    sheet["S81"].value = device.fields["data_description"].get()[0]
    sheet["S82"].value = device.fields["data_description"].get()[1]
    sheet["S83"].value = device.fields["data_description"].get()[2]
    sheet["S84"].value = device.fields["data_description"].get()[3]
    sheet["S85"].value = device.fields["data_description"].get()[4]
    # ---------------------------
    # GUARANTEES
    # ---------------------------
    sheet["S88"].value = device.guarantee.fields["dNLGuarantee"].get()
    sheet["S89"].value = device.guarantee.fields["dLLGuarantee"].get()
    sheet["S90"].value = f"{device.guarantee.fields['dIZGuarantee'].get()}%"
    sheet["S91"].value = device.guarantee.fields["dExcGuarantee"].get()
    # ---------------------------
    # GENERAL OPTIONALS & COMMENTS
    # ---------------------------
    sheet["S93"].value = device.fields["sConnectionDiagramsItems"].get()
    sheet["S94"].value = device.fields["dOverallRBFactor"].get()
    sheet["S95"].value = device.fields["dAltitude"].get()
    sheet["S96"].value = device.fields["sWindingSpecComments"].get()
    sheet["S97"].value = device.core.fields["sSheetComments"].get()
    sheet["S98"].value = device.fields["sPurchasingComments"].get()
    sheet["S99"].value = device.fields["sInsulationComments"].get()
    sheet["S100"].value = device.fields["dONANLosses"].get()
    sheet["S101"].value = device.fields["dONANTOR"].get()
    # ---------------------------
    # OPTIONAL CORE DATA
    # ---------------------------
    sheet["S103"].value = device.core.fields["dWindingHeightAdder"].get()
    sheet["S104"].value = device.core.fields["dWindingWidthAdder"].get()
    sheet["S105"].value = device.core.fields["dCoilCorePBEnd"].get()
    sheet["S106"].value = device.core.fields["dCoilCoilPBSides"].get()
    sheet["S107"].value = device.core.fields["dCoilCorePBSides"].get()
    sheet["S108"].value = device.core.fields["dFixedSmWindowWidth"].get()
    sheet["S109"].value = device.core.fields["dFixedWindowHeight"].get()
    # ---------------------------
    # OPTIONAL LV DATA
    # ---------------------------
    sheet["S112"].value = pWindL.fields["sDuctPlacement"].get()
    sheet["S113"].value = pWindL.fields["dLayerInsulation"].get()
    sheet["S114"].value = pWindL.fields["dOptStayback"].get()
    sheet["S115"].value = pWindL.fields["dBuildFactor"].get()
    sheet["S116"].value = pWindL.fields["bEndFiller"].get().value
    # ---------------------------
    # OPTIONAL HV DATA
    # ---------------------------
    sheet["S120"].value = pWindH.fields["sDuctPlacement"].get()
    sheet["S121"].value = pWindH.fields["dOptStayback"].get()
    sheet["S122"].value = pWindH.fields["dLayerInsulation"].get()
    sheet["S123"].value = pWindH.fields["dEndInsulation"].get()
    sheet["S124"].value = pWindH.fields["dLeadPadWitdh"].get()
    sheet["S125"].value = pWindH.fields["iLeadPads"].get()
    # ---------------------------
    # OPT BIL UNDER HV
    # ---------------------------
    if pWindH.fields["iBILUnder"].isDefault()==False:
        sheet["S126"].value = int(pWindH.fields["iBILUnder"].get().name.split("_")[-1])
    sheet["S127"].value = pWindH.fields["sHLUnder"].get()
    sheet["S128"].value = pWindH.fields["dHLUnderRB"].get()
    sheet["S129"].value = pWindH.fields["sHLOver"].get()
    sheet["S130"].value = pWindH.fields["dHLOverRB"].get()
    sheet["S131"].value = pWindH.fields["dBuildFactor"].get()
    sheet["S132"].value = pWindH.fields["bEndFiller"].get().value
    sheet["S133"].value = pWindH.fields["sPHTapLeads"].get()
    # ---------------------------
    # ROUTINE TESTS
    # ---------------------------
    sheet["S136"].value = device.tests.fields["tHipot"].get().value
    sheet["S137"].value = device.tests.fields["dHipotHV"].get()
    sheet["S138"].value = device.tests.fields["dHipotLV"].get()
    sheet["S139"].value = device.tests.fields["bInduced"].get().value
    sheet["S140"].value = device.tests.fields["dXRatedVoltage"].get()
    sheet["S141"].value = device.tests.fields["bResistance"].get().value
    sheet["S142"].value = device.tests.fields["bOtherResistance"].get().value
    sheet["S143"].value = "" if device.tests.fields["tTapsExtremesORAII"].get()==eTapExtremes.NONE else device.tests.fields["tTapsExtremesORAII"].get().value
    sheet["S144"].value = device.tests.fields["sSampleRate"].get()
    sheet["S145"].value = device.tests.fields["bCoreLossTest"].get().value
    sheet["S146"].value = device.tests.fields["dExtraNoLoad"].get()
    sheet["S147"].value = device.tests.fields["bLoadLossTest"].get().value
    sheet["S148"].value = device.tests.fields["bLoadLossTap"].get().value
    sheet["S149"].value = device.tests.fields["d3LoadLosskVA"].get()
    sheet["S150"].value = device.tests.fields["d3LoadLossTap"].get()
    sheet["S151"].value = device.tests.fields["d4LoadLosskVA"].get()
    sheet["S152"].value = device.tests.fields["d4LoadLossTap"].get()
    sheet["S153"].value = device.tests.fields["sOtherRoutineComments"].get()
    # ---------------------------
    # TYPE TESTS
    # ---------------------------
    sheet["S156"].value = device.tests.fields["bTemperatureRise"].get().value
    if device.tests.fields["tSpecificationsTR"].get()==eStandardTemperatureRise.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsTR"].get()==eStandardTemperatureRise.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsTR"].get().value
    sheet["S157"].value = result
    sheet["S158"].value = device.tests.fields["sCommentsTR"].get()
    sheet["S159"].value = device.tests.fields["bLightningImpulse"].get().value
    if device.tests.fields["tSpecificationsLI"].get()==eStandardLightningImpulse.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsLI"].get()==eStandardLightningImpulse.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsLI"].get().value
    sheet["S160"].value = result
    sheet["S161"].value = device.tests.fields["tSequenceLI"].get().value
    sheet["S162"].value = device.tests.fields["tImpulse"].get().value
    sheet["S163"].value = device.tests.fields["sCommentsLI"].get()
    sheet["S164"].value = device.tests.fields["bRIV"].get().value
    if device.tests.fields["tSpecificationsRIV"].get()==eStandardRIV.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsRIV"].get()==eStandardRIV.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsRIV"].get().value
    sheet["S165"].value = result
    sheet["S166"].value = device.tests.fields["sCommentsRIV"].get()
    sheet["S167"].value = device.tests.fields["bSoundLevel"].get().value
    if device.tests.fields["tSpecificationsSL"].get()==eStandardSoundLevel.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsSL"].get()==eStandardSoundLevel.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsSL"].get().value
    sheet["S168"].value = device.tests.fields["tSpecificationsSL"].get().value
    sheet["S169"].value = device.tests.fields["sCommentsSL"].get()
    sheet["S170"].value = device.tests.fields["sOtherRoutineCommentsType"].get()
    # ---------------------------
    # OPTIONAL FORM DATA
    # ---------------------------
    sheet["S172"].value = device.fields["sFormMaterial"].get()
    sheet["S173"].value = device.fields["dFormMaterialRB"].get()

    sheet["F91"].value = pCoilL.fields["iSectionsLeg"].get()
    sheet["I91"].value = pCoilH.fields["iSectionsLeg"].get()

def three_phases(sheet,device):

    #sheet["C81"].value = device.fields["array_type"].get().name
    sheet["P90"].value = device.fields["so_number"].get()

    sheet["P91"].value = device.fields["customer_name"].get()
    sheet["P92"].value = device.fields["customer_code"].get()

    sheet["P93"].value = device.guarantee.fields["dEvaluation_nl"].get()
    sheet["P94"].value = device.guarantee.fields["dEvaluation_ll"].get()
    sheet["P95"].value = device.fields["issued_by"].get()
    date_str = device.fields["original_issue_date"].get()  # '11/18/2025'

    date_obj = datetime.strptime(date_str, "%m/%d/%Y")

    sheet["P96"].value = date_obj

    sheet["P97"].value = device.guarantee.fields["sCustomerSpec"].get()
    sheet["P98"].value = device.guarantee.fields["tDesignSpec"].get().value
    sheet["P99"].value = device.guarantee.fields["tEfficiencySpec"].get().value
    if device.guarantee.fields["bNLCorrected_to_85C"].get()==eLogical.TRUE:
        sheet["P100"].value = "Yes"
    else:
        sheet["P100"].value = "No"
    sheet["P101"].value = device.guarantee.fields["dNL_pu"].get()
    sheet["P102"].value = device.guarantee.fields["dGroundShortTime"].get()
    sheet["P103"].value = device.guarantee.fields["dVirtualZ"].get()

    #General Specs
    sheet["P105"].value = device.fields["nominal_kva"].get()

    tCoolingTypes = device.cooling.fields["tCoolingTypes"].get()
    if tCoolingTypes==eCoolingType.COOL_ONAN:
        sheet["P106"].value="ONAN"
    elif tCoolingTypes==eCoolingType.COOL_ONAN_ONAF:
        sheet["P106"].value="ONAN/ONAF"
    elif tCoolingTypes==eCoolingType.COOL_ONAN_ONAF_ONAF:
        sheet["P106"].value="ONAN/ONAF/ONAF"
    elif tCoolingTypes==eCoolingType.COOL_KNAN:
        sheet["P106"].value="KNAN"
    elif tCoolingTypes==eCoolingType.COOL_KNAN_KNAF:
        sheet["P106"].value="KNAN/KNAF"
    elif tCoolingTypes==eCoolingType.COOL_KNAN_KNAF_KNAF:
        sheet["P106"].value="KNAN/KNAF/KNAF"
    
    sheet["P107"].value = device.fields["dPerSpcFor"].get()

    #P108 falta
    if device.cooling.fields["tProvONAF"].get()==eLogical.TRUE: 
        sheet["P108"].value="Yes"  
    else:
        sheet["P108"].value="No"
    sheet["P109"].value = device.fields["frequency"].get()
    tTemperatureRises = device.cooling.fields["tTemperatureRises"].get()
    if tTemperatureRises==eTemperatureRise.Temp_55_65: 
        sheet["P110"].value=55 
        sheet["P111"].value="Yes" 
    else:
        sheet["P111"].value="No"
        if tTemperatureRises==eTemperatureRise.Temp_65:
            sheet["P110"].value=65 
        elif tTemperatureRises==eTemperatureRise.Temp_55:
            sheet["P110"].value=55 

    sheet["P112"].value = device.fields["transformer_family"].get().name

    sheet["P113"].value = device.fields["winding_description"].get()

    pWindH = device.getHVWind()
    pWindL = device.getLVWind()

    #Taps
    # iNumberPositions = sheet["P118"].value
    # iPositionsAboveRated = sheet["P117"].value
    # if iNumberPositions>0 and iPositionsAboveRated>0:
    #     iPasos = (iNumberPositions-1)/iPositionsAboveRated 
    if pWindH.fields["bHaveTaps"].get()==eLogical.TRUE:
        sheet["P115"].value="Yes"
        changer = pWindH.changer
        sheet["P116"].value = float(changer.fields["dPerUp"].get()/changer.taps_sup())  # Porcentaje paso
        sheet["P117"].value = pWindH.fields["iNomPos"].get()  # nominal
        sheet["P118"].value = changer.get_NumSteps()+1   # Número de posiciones
        sheet["P119"].value = changer.fields["sPerTaps"].get()
        sheet["P120"].value = changer.fields["tCapacity"].get().value 
    else:
        sheet["P115"].value="No"

    #Voltages, Connections, BIL
    sheet["P122"].value = pWindH.fields["dVoltage"].get()
    sheet["P123"].value = 1.73 if pWindH.fields["type_connection"].get()==eConnectionType.DELTA else 1

    sheet["P124"].value = int(pWindH.fields["iBIL"].get().name.split("_")[-1])
    
    sheet["P125"].value = "Yes" if pWindH.fields["ground_Y"].get()==eLogical.TRUE else "No"

    sheet["P129"].value = pWindL.fields["dVoltage"].get()
    sheet["P130"].value = 1.73 if pWindL.fields["type_connection"].get()==eConnectionType.DELTA else 1
    sheet["P131"].value = int(pWindL.fields["iBIL"].get().name.split("_")[-1])

    #LV Leads & Bus  
    sheet["P133"].value = pWindL.fields["iQuantityLeads"].get()
    sheet["P134"].value = pWindL.fields["dThicknessLeads"].get()
    sheet["P135"].value = pWindL.fields["dWidthLeads"].get()
    sheet["P136"].value = pWindL.fields["tMaterialLeads"].get().name

    sheet["P137"].value = device.fields["iQuantityBuswork"].get()
    sheet["P138"].value = device.fields["dThicknessBuswork"].get()
    sheet["P139"].value = device.fields["dWidthBuswork"].get()
    sheet["P140"].value = device.fields["tMaterialBuswork"].get().name
    
    sheet["P141"].value = device.mechanical.fields["d_kW2"].get()[1]*1000#device.cooling.fields["dMaxONAF"].get()[1]
    sheet["P142"].value = device.cooling.fields["dTORONAF"].get()[1]#device.cooling.fields["dTORONAF"].get()[1]
    sheet["P143"].value = device.cooling.fields["tTemperatureRise"].get()[1].value
    #sheet["P144"].value = device.cooling.fields["dMaxONAF"].get()[2]
    #sheet["P145"].value = device.cooling.fields["dTORONAF"].get()[2]

    #Core
    core = device.core
    sheet["P148"].value = "Yes" if core.fields["tIsolatedCore"].get()==eLogical.TRUE else "No"
    sheet["P149"].value = core.fields["tLaminationType"].get().value
    

    sheet["P150"].value = core.fields["dTotalStack"].get()
    sheet["P151"].value = core.fields["dSteelWidth"].get()
    sheet["P152"].value = core.fields["dPerOvervoltageMaxFlux"].get()
    #get_cell_value(device.core.fields["dTestInductionLevel"],sheet,"P140")#Test Induction Level
    sheet["P153"].value = core.fields["dMaxWattsPound"].get()
    sheet["P154"].value = core.fields["dInductionLevel"].get()
    sheet["P155"].value = core.fields["iNoLoops"].get()
    sheet["P156"].value = core.fields["dLossFact"].get()

    #LV winding
    sheet["P158"].value = pWindL.fields["dNomTurnsUI"].get()
    sheet["P159"].value = pWindL.coils[0].fields["iLayers"].get()
    if pWindL.coils[0].conductor[0].fields["iType"].get()==eTypeConductor.STRIP:
       sheet["P160"].value=1
       sheet["P162"].value = pWindL.coils[0].conductor[0].fields["dWidth"].get()
    else:
       sheet["P160"].value=2
    sheet["P161"].value = pWindL.coils[0].conductor[0].fields["iNoConductor"].get()

    sheet["P163"].value = pWindL.coils[0].conductor[0].fields["iNoAxial"].get()
    sheet["P164"].value =  pWindL.coils[0].conductor[0].fields["iNoRadial"].get()
    #P167 falta
    sheet["P168"].value = pWindL.fields["dLeadBuild"].get()
    sheet["P169"].value = pWindL.fields["iNoDuctInner"].get()
    sheet["P170"].value = pWindL.fields["iNoDuctOuter"].get()
    sheet["P171"].value = pWindL.fields["dDuctSize"].get()
    if pWindL.fields["tCoating"].get()==eCoating.ENAMEL: 
        sheet["P172"].value="Enamel"
    elif pWindL.fields["tCoating"].get()==eCoating.PAPER:
        sheet["P172"].value="Paper"
    else:
        sheet["P172"].value=""

    #HV winding
    sheet["P174"].value = "Yes" if pWindH.coils[0].fields["bConstantStayback"].get()==eLogical.TRUE else "No"
    if pWindH.coils[0].conductor[0].fields["iType"].get()==eTypeConductor.STRIP:
        sheet["P175"].value = 1
        sheet["P177"].value = pWindH.coils[0].conductor[0].fields["dWidth"].get()
    else:
        sheet["P175"].value = 2

    sheet["P176"].value = pWindH.coils[0].conductor[0].fields["iNoConductor"].get()
    sheet["P178"].value = pWindH.coils[0].conductor[0].fields["iNoAxial"].get()
    sheet["P179"].value = pWindH.coils[0].conductor[0].fields["iNoRadial"].get()
    sheet["P183"].value = pWindH.coils[0].fields["iLayers"].get()
    sheet["P184"].value = "Yes" if pWindH.fields["bHalfTurns"].get()==eLogical.TRUE else "No"
    sheet["P185"].value = pWindH.fields["iNoDuctInner"].get()
    #P186 falta
    sheet["P187"].value = pWindH.fields["sPHTapLeads"].get()
    #get_cell_value(pWindH.fields["dCoilLineLeads"],sheet,"P187")
    sheet["P188"].value = pWindH.fields["dLeadBuild"].get()
    if pWindH.have_taps():
        if pWindH.changer.fields["tModelChanger"].get()!=eModelChanger.NONE:
            sheet["P189"].value = pWindH.changer.fields["tModelChanger"].get().value

    #Revision data
    sheet["S91"].value = device.fields["data_description"].get()[0] 
    sheet["S92"].value = device.fields["data_description"].get()[1] 
    sheet["S93"].value = device.fields["data_description"].get()[2] 
    sheet["S94"].value = device.fields["data_description"].get()[3] 
    sheet["S95"].value = device.fields["data_description"].get()[4] 

    #Guarantees
    sheet["S98"].value = device.guarantee.fields["dNLGuarantee"].get()
    sheet["S99"].value = device.guarantee.fields["dLLGuarantee"].get()
    #IZ Guarantee
    sheet["S100"].value = f"Min. {device.guarantee.fields['dIZGuarantee'].get()}%"
    sheet["S101"].value = device.guarantee.fields["dExcGuarantee"].get()

    #General Optionals & Comments
    sheet["S103"].value = device.fields["sConnectionDiagramsItems"].get()
    sheet["S104"].value = device.fields["dOverallRBFactor"].get()
    sheet["S105"].value = device.fields["dAltitude"].get()
    sheet["S106"].value = device.fields["sWindingSpecComments"].get()
    sheet["S108"].value = device.fields["sPurchasingComments"].get()
    sheet["S109"].value = device.fields["sInsulationComments"].get()
    sheet["S110"].value = device.fields["dONANLosses"].get()
    sheet["S111"].value = device.fields["dONANTOR"].get()

    #Optional Core Data
    sheet["S113"].value = device.core.fields["dWindingHeightAdder"].get()
    sheet["S114"].value = device.core.fields["dWindingWidthAdder"].get()
    sheet["S115"].value = device.core.fields["dCoilCorePBEnd"].get()
    sheet["S116"].value = device.core.fields["dCoilCoilPBSides"].get()
    sheet["S117"].value = device.core.fields["dCoilCorePBSides"].get()
    sheet["S118"].value = device.core.fields["dFixedSmWindowWidth"].get()
    sheet["S119"].value = device.core.fields["dFixedWindowHeight"].get()
    sheet["S120"].value = device.core.fields["dFixedWindowLarge"].get()

     #Optional Form Data
    sheet["S123"].value = device.fields["sFormMaterial"].get()
    sheet["S123"].value = device.fields["dFormMaterialRB"].get()
    sheet["S124"].value = device.fields["sFormComments"].get()

    #Optional LV Data
    sheet["S126"].value = pWindL.fields["sDuctPlacement"].get()
    sheet["S127"].value = pWindL.fields["dLayerInsulation"].get()
    sheet["S128"].value = pWindL.fields["dOptStayback"].get()
    sheet["S129"].value = pWindL.fields["dBuildFactor"].get()
    if pWindL.fields["bEndfillerOpt"].get()==eLogical.TRUE:
        sheet["S130"].value = "Yes"

    sheet["S131"].value = pWindL.fields["iLeadPads"].get()
    sheet["S132"].value = pWindL.fields["dNoseAdder"].get()

    #Optional HV Data
    sheet["S134"].value = pWindH.fields["sDuctPlacement"].get()
    sheet["S135"].value = pWindH.fields["dOptStayback"].get()
    sheet["S136"].value = pWindH.fields["dLayerInsulation"].get()
    sheet["S137"].value = pWindH.fields["dEndInsulation"].get() 
    sheet["S138"].value = pWindH.fields["dLeadPadWitdh"].get()
    sheet["S139"].value = pWindH.fields["iLeadPads"].get()
    #Opt BIL Under HV
    if pWindH.fields["iBILUnder"].isDefault()==False:
        sheet["S140"].value = int(pWindH.fields["iBILUnder"].get().name.split("_")[-1])
    
    sheet["S141"].value = pWindH.fields["sHLUnder"].get()
    sheet["S142"].value = pWindH.fields["dHLUnderRB"].get()
    sheet["S143"].value = pWindH.fields["sHLOver"].get()
    sheet["S144"].value = pWindH.fields["dHLOverRB"].get()
    sheet["S145"].value = pWindH.fields["dBuildFactor"].get()
    if pWindH.fields["bEndfillerOpt"].get()==eLogical.TRUE:
        sheet["S146"].value = "Yes"

    sheet["S147"].value = pWindH.fields["dNoseAdder"].get()
    #get_cell_value(pWindH.fields["sPHTapLeads"],sheet,"S133")#Optional HV Ph. / Tap Leads

    #Routine Tests
    sheet["S149"].value = device.tests.fields["tHipot"].get().value
     
    sheet["S150"].value = device.tests.fields["dHipotHV"].get()
    sheet["S151"].value = device.tests.fields["dHipotLV"].get()

    sheet["S152"].value = "Yes" if device.tests.fields["bInduced"].get()==eLogical.TRUE else "No"
    sheet["S153"].value = device.tests.fields["dXRatedVoltage"].get()

    sheet["S154"].value = "Yes" if device.tests.fields["bResistance"].get()==eLogical.TRUE else "No"
        
    sheet["S155"].value = "Yes" if device.tests.fields["bOtherResistance"].get()==eLogical.TRUE else "No"

    sheet["S156"].value = "" if device.tests.fields["tTapsExtremesORAII"].get()==eTapExtremes.NONE else device.tests.fields["tTapsExtremesORAII"].get().value
    sheet["S157"].value = device.tests.fields["sSampleRate"].get()
    sheet["S158"].value = device.tests.fields["bCoreLossTest"].get().value
    sheet["S159"].value = device.tests.fields["dExtraNoLoad"].get()
    sheet["S161"].value = device.tests.fields["bLoadLossTest"].get().value
    sheet["S162"].value = device.tests.fields["bLoadLossTap"].get().value
    sheet["S163"].value = device.tests.fields["d3LoadLosskVA"].get()
    sheet["S164"].value = device.tests.fields["d3LoadLossTap"].get()
    sheet["S165"].value = device.tests.fields["d4LoadLosskVA"].get()
    sheet["S166"].value = device.tests.fields["d4LoadLossTap"].get()
    sheet["S167"].value = device.tests.fields["d5LoadLosskVA"].get()
    sheet["S168"].value = device.tests.fields["d5LoadLossTap"].get()
    sheet["S169"].value = device.tests.fields["dDissipationFactor"].get()
    sheet["S171"].value = device.tests.fields["sOtherRoutineComments"].get()

    
    #Type Tests
    sheet["S173"].value = device.tests.fields["bTemperatureRise"].get().value
    if device.tests.fields["tSpecificationsTR"].get()==eStandardTemperatureRise.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsTR"].get()==eStandardTemperatureRise.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsTR"].get().value
    sheet["S174"].value = result
    sheet["S175"].value = device.tests.fields["sCommentsTR"].get()
    sheet["S176"].value = device.tests.fields["bLightningImpulse"].get().value
    if device.tests.fields["tSpecificationsLI"].get()==eStandardLightningImpulse.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsLI"].get()==eStandardLightningImpulse.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsLI"].get().value
    sheet["S177"].value = result
    sheet["S178"].value = device.tests.fields["tSequenceLI"].get().value
    sheet["S179"].value = device.tests.fields["tImpulse"].get().value
    
    sheet["S180"].value = device.tests.fields["sCommentsLI"].get()
    sheet["S181"].value = device.tests.fields["bRIV"].get().value
    if device.tests.fields["tSpecificationsRIV"].get()==eStandardRIV.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsRIV"].get()==eStandardRIV.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsRIV"].get().value
    sheet["S182"].value = result
    sheet["S183"].value = device.tests.fields["sCommentsRIV"].get()
    sheet["S184"].value = device.tests.fields["bSoundLevel"].get().value
    if device.tests.fields["tSpecificationsSL"].get()==eStandardSoundLevel.NONE:
        result = ""
    elif device.tests.fields["tSpecificationsSL"].get()==eStandardSoundLevel.IEEEC571290:
        result = "ANSI C57.12.90"
    else:
        result = device.tests.fields["tSpecificationsSL"].get().value
    sheet["S185"].value = device.tests.fields["tSpecificationsSL"].get().value
    sheet["S186"].value = device.tests.fields["sCommentsSL"].get()
    sheet["S187"].value = device.tests.fields["sOtherRoutineCommentsType"].get()
    
    # texto = sheet["F189"].value
    # resultado = False
    # if texto is not None:
    #     resultado = "per k4" in texto.lower()
    # if resultado:
    #     device.cooling.fields["tK4Factor"].set(eLogical.TRUE)
    sheet["F90"].value = pWindL.coils[0].fields["iSectionsLeg"].get()
    sheet["I90"].value = pWindH.coils[0].fields["iSectionsLeg"].get()

    # get_cell_value(device.core.fields["dLamFactor"],sheet,"Z21")
    # get_cell_value(device.core.fields["dStackTolPIn"],sheet,"Z22")
    # get_cell_value(device.core.fields["dStackTolMIn"],sheet,"Z23")