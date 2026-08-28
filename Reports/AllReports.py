from http.client import LOOP_DETECTED
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
from datetime import datetime
import win32com.client as win32
import os
from Resources.resources import resource_path
from Definitions.functions import address,ADDRESSZ,PI,is_folder_open
from datetime import datetime
from Reports.QuotePage import make_quote_page
from Reports.WindingSpecification import make_winding_specification
from Reports.EngineeringData import make_engineering_data
from Reports.Core_Drawing import make_core_drawing
from Definitions.enums import eLaminationType,eWarnings,eLogical
import shutil
from pathlib import Path

def make_excel_reports(device = None):

    def set_inside_border(ws,min_row, max_row, min_col, max_col, border = None, fillE = None):
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border  # apply same border to every cell
                if fillE is not None:
                    cell.fill = fillE


    def set_outer_border(ws, min_row, max_row, min_col, max_col, border, fillE = None):
        """
        Applies a border only around the perimeter of a cell range.

        ws : Worksheet
        min_row : starting row of the range
        max_row : ending row of the range
        min_col : starting column of the range
        max_col : ending column of the range
        border : Border object (e.g., Border(left=Side(...)))
        """
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                # Borde exterior superior
                top = border.top if row == min_row else cell.border.top
                # Borde exterior inferior
                bottom = border.bottom if row == max_row else cell.border.bottom
                # Borde exterior izquierdo
                left = border.left if col == min_col else cell.border.left
                # Borde exterior derecho
                right = border.right if col == max_col else cell.border.right
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
                if fillE is not None:
                    cell.fill = fillE

    def set_size_cells(ws, min_row, max_row, min_col, max_col, size):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = size

    def set_alignment_cells(ws, min_row, max_row, min_col, max_col, horizontal, vertical):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    sAddress = ADDRESSZ#address(device.fields["so_number"].get())
    #documents_folder = Path.home() / "Documents"
    so_number = str(device.fields["so_number"].get())  # ensure it's a string
    sAddress = os.path.join(sAddress, so_number)
    #sAddress = documents_folder / so_number
    os.makedirs(sAddress, exist_ok=True)  

    #Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    img_path = resource_path("Resources/logo_reports.png")
    img_pil = PILImage.open(img_path)

    dpi = 96
    width_inch = 1.2
    height_inch = 0.37
    width_px = int(width_inch * dpi)
    height_px = int(height_inch * dpi)

    img_pil = img_pil.resize((width_px, height_px), PILImage.Resampling.LANCZOS)

    img_bytes = BytesIO()
    img_pil.save(img_bytes, format="PNG")

    #img_bytes.seek(0)
    img_data = img_bytes.getvalue()

    img = Image(BytesIO(img_data))
    img2 = Image(BytesIO(img_data))
    img3 = Image(BytesIO(img_data))

    iNoPhases = device.fields["iNoPhases"].get()
    tLamination = device.core.fields["tLaminationType"].get()
    bAmorphous = True if tLamination == eLaminationType.Am or tLamination == eLaminationType.AmHB1 else False
    pWindH = device.getHVWind()
    bHaveTaps = pWindH.have_taps()
    pHCoil = pWindH.coils[0]
    ground_YHV = pWindH.fields["ground_Y"].get()

    if iNoPhases==1:
        if bAmorphous:
            img_path_core = resource_path("Resources/single_phase_core_Am.png")
        else:
            img_path_core = resource_path("Resources/single_phase_core_M0H.png")
    else:
        if bAmorphous:
            img_path_core = resource_path("Resources/three_phases_core_Am.png")
        else:
            img_path_core = resource_path("Resources/three_phases_core_M0H.png")
    img_pil_core = PILImage.open(img_path_core)

    dpi_core = 96
    if iNoPhases==1:
        width_inch_core = 4.22
        height_inch_core = 2.5
    else:
        width_inch_core = 6.22
        height_inch_core = 2.5
    width_px_core = int(width_inch_core * dpi_core)
    height_px_core = int(height_inch_core * dpi_core)

    img_pil_core = img_pil_core.resize((width_px_core, height_px_core), PILImage.Resampling.LANCZOS)

    img_bytes_core = BytesIO()
    img_pil_core.save(img_bytes_core, format="PNG")
    img_bytes_core.seek(0)

    img_core = Image(img_bytes_core)


    ws.page_margins = PageMargins(
                        left=0.25,
                        right=0.25,
                        top=0.25,
                        bottom=0.25,
                        header=0,
                        footer=0
                    )

    # Definir estilos
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))


    thick_border = Border(left=Side(style='medium'),   # también puedes usar 'thick'
                          right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    dotted_border = Border(
                            left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

    # Fondo de puntitos (pattern fill)
    dot_fill = PatternFill(
                            fill_type="gray125",  # patrón de puntos
                            start_color="000000", # color de los puntos: negro
                            end_color="FFFFFF"    # color de fondo: blanco (o transparente)
                        )
    
    dot_fill3 = PatternFill(
                            fill_type="gray125",  # patrón de puntos
                            start_color="FFFF0000", # color de los puntos: negro
                            end_color="FFFFFF"    # color de fondo: blanco (o transparente)
                        )
    
    white_fill = PatternFill(
                            fill_type="solid",
                            start_color="FFFFFF",
                            end_color="FFFFFF"
                        )

    yellow_fill = PatternFill(
                            fill_type="solid",
                            start_color="FFFF00",  # amarillo
                            end_color="FFFF00"
                        )

    helv_font11 = Font(name="Helvetica", size=11)
    helv_font = Font(name="Helvetica", size=10)
    helv_font_bold = Font(name="Helvetica", size=10, bold = True)
    helv_font_red = Font(name="Helvetica", size=10,color="FF0000")
    helv_font_bold_under = Font(name="Helvetica", size=10, bold = True,underline="single")
    helv_font9 = Font(name="Helvetica", size=9)
    helv_font9_bold = Font(name="Helvetica", size=9, bold = True)
    helv_font8 = Font(name="Helvetica", size=8)
    helv_font8_bold = Font(name="Helvetica", size=8, bold = True)
    Cal_font8 = Font(name="Calibri", size=8)

    # Opcional: orientación y ajuste a ancho/alto
    #ws.page_setup.orientation = 'portrait'  # 'landscape' también
    # Datos de la hoja (A1:L69)

    if iNoPhases==1:
        # Ajustar anchos de columna
        column_widths = {
            'A': 1.5, 
            'B': 6, 
            'C': 12.29, 
            'D': 10.14, 
            'E': 12.29, 
            'F': 12.29, 
            'G': 12.29, 
            'H': 12.29, 
            'I': 12.29, 
            'J': 14.86,
            'K': 12.29, 
            'L': 6, 
            'M': 0.92,
        }
    
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        #ws.print_area = "A1:L64"
        ws.page_setup.scale = 80
    
        for row in range(1, 500):
            ws.row_dimensions[row].height = 14.25  

        #QUOTE PAGE
        data = make_quote_page(device)

        for row in ws.iter_rows(min_row=1, max_row=272, min_col=1, max_col=13):
            for cell in row:
                cell.fill = white_fill

        for row in ws.iter_rows(min_row=5, max_row=6, min_col=3, max_col=3):
            for cell in row:
                cell.fill = dot_fill

        for row in ws.iter_rows(min_row=9, max_row=9, min_col=9, max_col=9):
            for cell in row:
                cell.fill = dot_fill

        set_inside_border(ws, min_row=3, max_row=6, min_col=4, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=3, max_row=4, min_col=3, max_col=3, border=dotted_border)
        set_inside_border(ws, min_row=8, max_row=16, min_col=3, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=17, max_row=22, min_col=5, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=16, max_row=17, min_col=3, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=8, max_row=13, min_col=9, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=36, max_row=41, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=43, max_row=47, min_col=3, max_col=10, border=dotted_border)
        set_inside_border(ws, min_row=49, max_row=50, min_col=3, max_col=10, border=dotted_border)
        set_inside_border(ws, min_row=52, max_row=56, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=58, max_row=62, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=64, max_row=67, min_col=3, max_col=7, border=dotted_border)

        set_outer_border(ws, min_row=1, max_row=68, min_col=2, max_col=12, border=thin_border)

        set_outer_border(ws, min_row=3, max_row=6, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=4, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=5, max_row=6, min_col=3, max_col=3, border=thin_border)    
        set_outer_border(ws, min_row=3, max_row=6, min_col=4, max_col=5, border=thin_border)  
        set_outer_border(ws, min_row=3, max_row=6, min_col=6, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=6, min_col=8, max_col=9, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=6, min_col=10, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=8, max_row=23, min_col=3, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=15, max_row=21, min_col=10, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=9, max_row=9, min_col=5, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=16, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=16, min_col=4, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=5, max_col=5, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=6, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=7, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=17, max_row=23, min_col=3, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=9, max_col=9, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=11, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=4, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=5, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=7, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=15, max_row=15, min_col=10, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=16, max_row=16, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=16, max_row=16, min_col=11, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=18, max_row=21, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=18, max_row=21, min_col=11, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=36, max_row=41, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=4, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=7, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=4, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=7, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=43, max_row=47, min_col=3, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=47, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=47, min_col=4, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=43, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=43, min_col=4, max_col=10, border=thin_border)

        set_outer_border(ws, min_row=49, max_row=50, min_col=3, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=50, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=49, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=49, min_col=4, max_col=10, border=thin_border)
    
        set_outer_border(ws, min_row=52, max_row=56, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=56, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=56, min_col=4, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=52, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=52, min_col=4, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=58, max_row=62, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=62, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=62, min_col=4, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=58, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=58, min_col=4, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=64, max_row=67, min_col=3, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=64, max_row=64, min_col=3, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=64, max_row=64, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=67, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=67, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=65, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=66, max_row=66, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=67, max_row=67, min_col=8, max_col=11, border=thin_border)

        row_iter = 0
        # Llenar la hoja con datos y aplicar formato
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_iter += 1
    
        ws.merge_cells("E8:F8")
        ws.merge_cells("I8:K8")
        ws.merge_cells("J15:K15")
        ws.merge_cells("J17:K17")
        ws.merge_cells("C18:D22")
        ws.merge_cells("D36:F36")
        ws.merge_cells("G36:K36")
        ws.merge_cells("D43:J43")
        ws.merge_cells("D52:K52")
        ws.merge_cells("D58:K58")
        ws.merge_cells("E64:G64")
        ws.merge_cells("E65:G65")
        ws.merge_cells("E66:G66")
        ws.merge_cells("E67:G67")
        ws.merge_cells("I65:K65")
        ws.merge_cells("I66:K66")
        ws.merge_cells("I67:K67")

        set_size_cells(ws, min_row=1, max_row=1, min_col=7, max_col=7, size=helv_font_bold_under)
        set_size_cells(ws, min_row=3, max_row=3, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=5, max_row=5, min_col=4, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=8, max_row=8, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=8, max_row=8, min_col=9, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=15, max_row=16, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=17, max_row=17, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=4, max_row=4, min_col=5, max_col=5, size=helv_font9)
        set_size_cells(ws, min_row=8, max_row=8, min_col=3, max_col=3, size=helv_font9_bold)
        set_size_cells(ws, min_row=37, max_row=37, min_col=6, max_col=6, size=helv_font8)
        set_size_cells(ws, min_row=43, max_row=43, min_col=3, max_col=3, size=helv_font9)
        set_size_cells(ws, min_row=50, max_row=50, min_col=3, max_col=3, size=helv_font9)

        d_values =[0.0,0.0]
        if device.losses.losses_warnings.warningValueCoreLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=10, max_row=10, min_col=11, max_col=11, size=helv_font_red)
        if device.losses.losses_warnings.warningValueLoadLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=11, max_row=11, min_col=11, max_col=11, size=helv_font_red)

        set_alignment_cells(ws, min_row=9, max_row=16, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=9, max_row=22, min_col=5, max_col=5, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=10, max_row=13, min_col=9, max_col=9, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=18, max_row=21, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=17, max_row=17, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=17, max_row=17, min_col=10, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=38, max_row=41, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=45, max_row=47, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=54, max_row=56, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=60, max_row=62, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=65, max_row=67, min_col=3, max_col=3, horizontal='left', vertical='center')

        

        ws.sheet_view.zoomScale = 70

        #winding specification
        data =  make_winding_specification(device)
        row_iter +=1 
        for row in ws.iter_rows(min_row = row_iter +2, max_row = row_iter +59, min_col=2, max_col=11):
            for cell in row:
                cell.fill = dot_fill

        # for row in ws.iter_rows(min_row = row_iter +2, max_row = row_iter +=3, min_col=2, max_col=4):
        #     for cell in row:
        #         cell.fill = white_fill

        set_inside_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=2, max_col=9, border=dotted_border)

        set_outer_border(ws, min_row = row_iter +2, max_row = row_iter +59, min_col=2, max_col=11, border=thin_border)
        set_outer_border(ws, min_row = row_iter +60, max_row = row_iter +63, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +67, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +2, max_row = row_iter +2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +3, max_row = row_iter +3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +2, max_row = row_iter +4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +2, max_row = row_iter +2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +3, max_row = row_iter +3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +4, max_row = row_iter +4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +6, max_row = row_iter +8, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +6, max_row = row_iter +8, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +6, max_row = row_iter +8, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +7, max_row = row_iter +8, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +10, max_row = row_iter +10, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +10, max_row = row_iter +10, min_col=6, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +11, max_row = row_iter +11, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +13, max_row = row_iter +15, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +13, max_row = row_iter +13, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +13, max_row = row_iter +13, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +14, max_row = row_iter +15, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +14, max_row = row_iter +15, min_col=8, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +16, max_row = row_iter +16, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +17, max_row = row_iter +17, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +18, max_row = row_iter +18, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +19, max_row = row_iter +19, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +20, max_row = row_iter +20, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +21, max_row = row_iter +21, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +22, max_row = row_iter +22, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +23, max_row = row_iter +23, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +24, max_row = row_iter +25, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +16, max_row = row_iter +16, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +17, max_row = row_iter +17, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +18, max_row = row_iter +18, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +19, max_row = row_iter +19, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +20, max_row = row_iter +20, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +21, max_row = row_iter +21, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +22, max_row = row_iter +22, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +23, max_row = row_iter +23, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +24, max_row = row_iter +25, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +16, max_row = row_iter +16, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +17, max_row = row_iter +17, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +18, max_row = row_iter +18, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +19, max_row = row_iter +19, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +20, max_row = row_iter +20, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +21, max_row = row_iter +21, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +22, max_row = row_iter +22, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +23, max_row = row_iter +23, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +24, max_row = row_iter +25, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +19, max_row = row_iter +19, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +22, max_row = row_iter +23, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +26, max_row = row_iter +27, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +26, max_row = row_iter +27, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +26, max_row = row_iter +27, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +26, max_row = row_iter +26, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +28, max_row = row_iter +28, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +29, max_row = row_iter +29, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +30, max_row = row_iter +30, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +31, max_row = row_iter +31, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +28, max_row = row_iter +28, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +29, max_row = row_iter +29, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +30, max_row = row_iter +30, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +31, max_row = row_iter +31, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +28, max_row = row_iter +28, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +29, max_row = row_iter +29, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +30, max_row = row_iter +30, min_col=8, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +33, max_row = row_iter +33, min_col=3, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +34, max_row = row_iter +35, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +37, max_row = row_iter +37, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=9, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +41, max_row = row_iter +41, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +41, max_row = row_iter +42, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +45, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +48, max_row = row_iter +51, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +48, max_row = row_iter +48, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +49, max_row = row_iter +51, min_col=8, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +49, max_row = row_iter +49, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +49, max_row = row_iter +49, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +58, max_row = row_iter +58, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +53, max_row = row_iter +53, min_col=3, max_col=9, border=thick_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=2, max_col=2, border=thick_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=3, max_col=9, border=thick_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +54, max_row = row_iter +55, min_col=10, max_col=10, border=thick_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +56, max_row = row_iter +58, min_col=11, max_col=11, border=thick_border, fillE = white_fill)

        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +65, max_row = row_iter +67, min_col=2, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +65, max_row = row_iter +67, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +65, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +64, max_row = row_iter +65, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        row_iter2 = 0

        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row_iter2 +=1

        s = f"F{2+row_iter}:H{2+row_iter}"
        ws.merge_cells(s)
        s = f"F{3+row_iter}:H{3+row_iter}"
        ws.merge_cells(s)
        s = f"F{4+row_iter}:H{4+row_iter}"
        ws.merge_cells(s)
        s = f"C{13+row_iter}:D{13+row_iter}"
        ws.merge_cells(s)
        s = f"C{15+row_iter}:D{15+row_iter}"
        ws.merge_cells(s)
        s = f"E{13+row_iter}:G{13+row_iter}"
        ws.merge_cells(s)
        s = f"H{13+row_iter}:J{13+row_iter}"
        ws.merge_cells(s)
        s = f"H{23+row_iter}:I{23+row_iter}"
        ws.merge_cells(s)
        s = f"C{33+row_iter}:J{33+row_iter}"
        ws.merge_cells(s)
        s = f"D{48+row_iter}:F{48+row_iter}"
        ws.merge_cells(s)
        s = f"I{49+row_iter}:K{49+row_iter}"
        ws.merge_cells(s)
        s = f"C{53+row_iter}:I{53+row_iter}"
        ws.merge_cells(s)
        s = f"C{64+row_iter}:E{64+row_iter}"
        ws.merge_cells(s)
        s = f"H{50+row_iter}:I{50+row_iter}"
        ws.merge_cells(s)
        s = f"H{51+row_iter}:I{51+row_iter}"
        ws.merge_cells(s)
        s = f"J{50+row_iter}:K{50+row_iter}"
        ws.merge_cells(s)
        s = f"C{65+row_iter}:E{65+row_iter}"
        ws.merge_cells(s)
        s = f"C{66+row_iter}:E{66+row_iter}"
        ws.merge_cells(s)
        s = f"C{67+row_iter}:E{67+row_iter}"
        ws.merge_cells(s)
        s = f"J{4+row_iter}:K{4+row_iter}"
        ws.merge_cells(s)

        #set_size_cells(ws, min_row = row_iter +3, max_row = row_iter +=3, min_col=11, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +7, max_row = row_iter +7, min_col=8, max_col=8, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +8, max_row = row_iter +8, min_col=6, max_col=6, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +10, max_row = row_iter +10, min_col=9, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +31, max_row = row_iter +31, min_col=3, max_col=4, size=helv_font8)
        set_size_cells(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=2, max_col=6, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +66, max_row = row_iter +66, min_col=7, max_col=8, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +54, max_row = row_iter +58, min_col=2, max_col=9, size=helv_font11)
        set_size_cells(ws, min_row = row_iter +54, max_row = row_iter +55, min_col=10, max_col=10, size=helv_font11)
        set_size_cells(ws, min_row = row_iter +57, max_row = row_iter +57, min_col=11, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row = row_iter +41, max_row = row_iter +41, min_col=11, max_col=11, size=helv_font8)

        set_size_cells(ws, min_row = row_iter +2, max_row = row_iter +2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +4, max_row = row_iter +4, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +10, max_row = row_iter +10, min_col=5, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +11, max_row = row_iter +11, min_col=3, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +13, max_row = row_iter +13, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +33, max_row = row_iter +33, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +48, max_row = row_iter +48, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +49, max_row = row_iter +49, min_col=8, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +60, max_row = row_iter +60, min_col=2, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=2, max_col=6, size=helv_font9_bold)
        set_size_cells(ws, min_row = row_iter +65, max_row = row_iter +65, min_col=6, max_col=6, size=helv_font8)
        set_size_cells(ws, min_row = row_iter +66, max_row = row_iter +66, min_col=6, max_col=6, size=helv_font8)
        set_size_cells(ws, min_row = row_iter +67, max_row = row_iter +67, min_col=6, max_col=6, size=helv_font8)

        set_alignment_cells(ws, min_row = row_iter +2, max_row = row_iter +3, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +2, max_row = row_iter +4, min_col=6, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +2, max_row = row_iter +4, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +6, max_row = row_iter +8, min_col=3, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +10, max_row = row_iter +10, min_col=10, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +15, max_row = row_iter +15, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +13, max_row = row_iter +15, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +13, max_row = row_iter +13, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +16, max_row = row_iter +24, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +19, max_row = row_iter +19, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +22, max_row = row_iter +23, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +25, max_row = row_iter +25, min_col=10, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +27, max_row = row_iter +27, min_col=3, max_col=3, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +26, max_row = row_iter +27, min_col=5, max_col=5, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +26, max_row = row_iter +26, min_col=8, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +28, max_row = row_iter +31, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +33, max_row = row_iter +35, min_col=3, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +38, max_row = row_iter +39, min_col=4, max_col=9, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +42, max_row = row_iter +43, min_col=4, max_col=9, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +41, max_row = row_iter +42, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +45, max_row = row_iter +46, min_col=4, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +48, max_row = row_iter +48, min_col=4, max_col=6, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=3, max_col=6, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +49, max_row = row_iter +49, min_col=9, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +50, max_row = row_iter +51, min_col=8, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +53, max_row = row_iter +58, min_col=2, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +64, max_row = row_iter +67, min_col=2, max_col=6, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +64, max_row = row_iter +64, min_col=8, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=7, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +66, max_row = row_iter +67, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row = row_iter +65, max_row = row_iter +67, min_col=3, max_col=5, horizontal='left', vertical='center')

        s = f"J{row_iter+64}"
        ws.add_image(img, s)

        row_iter += row_iter2+1

        data =  make_engineering_data(device)
    
        for row in ws.iter_rows(min_row=row_iter+2, max_row=row_iter+66, min_col=2, max_col=11):
            for cell in row:
                cell.fill = white_fill

        set_inside_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+10, max_row=row_iter+12, min_col=3, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+20, max_row=row_iter+22, min_col=3, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+32, max_row=row_iter+38, min_col=4, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+32, max_row=row_iter+41, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=4, max_col=9, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+49, max_row=row_iter+49, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+55, max_row=row_iter+59, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=5, max_col=6, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=7, max_col=8, border=dotted_border)

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+66, min_col=2, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+11, max_row=row_iter+11, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=3, max_col=5, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+16, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+16, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+16, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+16, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+16, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+16, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+17, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+16, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+16, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=8, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+18, max_row=row_iter+22, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+22, min_col=3, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+22, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+22, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=5, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+24, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+25, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=5, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+30, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+29, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+38, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+38, min_col=4, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+33, min_col=4, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=3, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+37, max_row=row_iter+37, min_col=3, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+38, max_row=row_iter+38, min_col=3, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+42, min_col=3, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+43, max_row=row_iter+43, min_col=4, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+41, min_col=10, max_col=11, border=dotted_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+42, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+41, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+41, min_col=11, max_col=11, border=thin_border, fillE = white_fill)    

        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+51, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+45, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+51, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+47, min_col=4, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+47, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+47, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+46, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+47, max_row=row_iter+49, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+50, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+51, max_row=row_iter+52, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+47, min_col=3, max_col=3, border=thin_border, fillE = dot_fill)

        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+59, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+61, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+61, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+56, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+61, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+61, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+66, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+66, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+64, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+65, max_row=row_iter+66, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+65, max_row=row_iter+66, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+64, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+65, max_row=row_iter+66, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+65, max_row=row_iter+66, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        row_iter2 = 0
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_iter2 +=1
    
        s = f"F{2+row_iter}:H{2+row_iter}"
        ws.merge_cells(s)
        s = f"F{3+row_iter}:H{3+row_iter}"
        ws.merge_cells(s)
        s = f"F{4+row_iter}:H{4+row_iter}"
        ws.merge_cells(s)
        s = f"C{14+row_iter}:F{14+row_iter}"
        ws.merge_cells(s)
        s = f"I{14+row_iter}:J{14+row_iter}"
        ws.merge_cells(s)
        s = f"C{18+row_iter}:I{18+row_iter}"
        ws.merge_cells(s)
        s = f"C{19+row_iter}:E{19+row_iter}"
        ws.merge_cells(s)
        s = f"D{20+row_iter}:E{20+row_iter}"
        ws.merge_cells(s)
        s = f"D{21+row_iter}:E{21+row_iter}"
        ws.merge_cells(s)
        s = f"D{22+row_iter}:E{22+row_iter}"
        ws.merge_cells(s)
        s = f"D{32+row_iter}:H{32+row_iter}"
        ws.merge_cells(s)
        s = f"D{33+row_iter}:F{33+row_iter}"
        ws.merge_cells(s)
        s = f"G{33+row_iter}:H{33+row_iter}"
        ws.merge_cells(s)
        s = f"J{32+row_iter}:K{32+row_iter}"
        ws.merge_cells(s)
        s = f"C{40+row_iter}:F{40+row_iter}"
        ws.merge_cells(s)
        s = f"G{40+row_iter}:H{40+row_iter}"
        ws.merge_cells(s)
        s = f"C{45+row_iter}:k{45+row_iter}"
        ws.merge_cells(s)
        s = f"D{46+row_iter}:E{46+row_iter}"
        ws.merge_cells(s)
        s = f"F{46+row_iter}:G{46+row_iter}"
        ws.merge_cells(s)
        ws.merge_cells(f"H{46+row_iter}:I{46+row_iter}")
        ws.merge_cells(f"J{46+row_iter}:K{46+row_iter}")
        ws.merge_cells(f"J{51+row_iter}:K{51+row_iter}")
        ws.merge_cells(f"J{52+row_iter}:K{52+row_iter}")
        ws.merge_cells(f"C{54+row_iter}:D{54+row_iter}")
        ws.merge_cells(f"E{54+row_iter}:F{54+row_iter}")
        ws.merge_cells(f"G{54+row_iter}:H{54+row_iter}")
        ws.merge_cells(f"J{54+row_iter}:K{54+row_iter}")
        ws.merge_cells(f"J{55+row_iter}:K{55+row_iter}")
        ws.merge_cells(f"J{56+row_iter}:K{56+row_iter}")
        ws.merge_cells(f"J{58+row_iter}:K{58+row_iter}")
        ws.merge_cells(f"J{61+row_iter}:K{61+row_iter}")
        ws.merge_cells(f"C{63+row_iter}:E{63+row_iter}")
        ws.merge_cells(f"C{64+row_iter}:E{64+row_iter}")
        ws.merge_cells(f"C{65+row_iter}:E{65+row_iter}")
        ws.merge_cells(f"I{63+row_iter}:K{64+row_iter}")
        ws.merge_cells(f"J{4+row_iter}:K{4+row_iter}")

        set_size_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=8, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=9, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=3, max_col=9, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=3, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+24, max_row=row_iter+24, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=4, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=7, max_col=8, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=7, max_col=7, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=8, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+45, max_row=row_iter+45, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=5, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=3, max_col=4, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=5, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=7, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=2, max_col=6, size=helv_font_bold)

        set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=6, max_col=6, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+7, max_row=row_iter+7, min_col=8, max_col=8, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+33, max_row=row_iter+33, min_col=7, max_col=7, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=6, max_col=6, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+64, max_row=row_iter+66, min_col=6, max_col=6, size=helv_font8)

        d_values = [0.0,0.0]
        if device.gradients.gradient_warnings.warningValueGradient(0,d_values) == eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=4, max_col=4, size=helv_font_red)
        if device.gradients.gradient_warnings.warningValueGradient(1,d_values) == eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=5, max_col=5, size=helv_font_red)
        if device.gradients.gradient_warnings.warningValueGradient(0,d_values) == eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=6, max_col=6, size=helv_font_red)
        if device.losses.losses_warnings.warningValueCoreLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+36, max_row=row_iter+36, min_col=4, max_col=4, size=helv_font_red)
        if device.losses.losses_warnings.warningValueLoadLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+36, max_row=row_iter+36, min_col=5, max_col=5, size=helv_font_red)
        if device.core.core_warnings.warningFluxDensity(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+22, max_row=row_iter+22, min_col=4, max_col=5, size=helv_font_red)
        
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=2, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=2, max_col=2, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+33, max_row=row_iter+42, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+33, max_row=row_iter+42, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+48, max_row=row_iter+51, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+47, max_row=row_iter+50, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+47, max_row=row_iter+50, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=5, max_col=5, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+59, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+59, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=5, max_col=5, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=6, max_col=6, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=7, max_col=7, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+55, max_row=row_iter+61, min_col=8, max_col=8, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+63, max_row=row_iter+63, min_col=7, max_col=7, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=3, max_col=5, horizontal='left', vertical='center')
        # Aplicar formato específico a celdas que necesitan diferente alineación
        # Celdas con texto largo - alineación izquierda
        
        s =f"J{63+row_iter}"
        ws.add_image(img2, s)

        row_iter += row_iter2+3
    
        #CORE DRAWING
        data = make_core_drawing(device)

        for row in ws.iter_rows(min_row=row_iter+1, max_row=row_iter+63, min_col=2, max_col=11):
            for cell in row:
                cell.fill = white_fill

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+62, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+18, min_col=4, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=6, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+23, max_row=row_iter+23, min_col=3, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+24, max_row=row_iter+26, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+30, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+31, max_row=row_iter+31, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+32, min_col=8, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=4, max_col=4, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=7, max_col=7, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+50, min_col=3, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+56, min_col=3, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+62, min_col=2, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+62, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+60, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+62, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+62, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+60, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+62, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+62, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        if bAmorphous:
            set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+42, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
            set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+42, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
            set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+41, min_col=4, max_col=6, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells(f"F{2+row_iter}:H{2+row_iter}")
        ws.merge_cells(f"F{20+row_iter}:H{20+row_iter}")
        ws.merge_cells(f"C{23+row_iter}:J{23+row_iter}")
        ws.merge_cells(f"H{28+row_iter}:J{28+row_iter}")
        ws.merge_cells(f"H{29+row_iter}:J{29+row_iter}")
        ws.merge_cells(f"H{30+row_iter}:J{30+row_iter}")
        ws.merge_cells(f"H{31+row_iter}:J{31+row_iter}")
        ws.merge_cells(f"H{32+row_iter}:J{32+row_iter}")
        ws.merge_cells(f"C{34+row_iter}:D{34+row_iter}")
        ws.merge_cells(f"C{35+row_iter}:D{35+row_iter}")
        ws.merge_cells(f"F{34+row_iter}:G{34+row_iter}")
        ws.merge_cells(f"F{35+row_iter}:G{35+row_iter}")
        ws.merge_cells(f"I{35+row_iter}:J{35+row_iter}")
        ws.merge_cells(f"C{59+row_iter}:E{59+row_iter}")
        ws.merge_cells(f"I{59+row_iter}:K{60+row_iter}")
        if bAmorphous:
            ws.merge_cells(f"E{40+row_iter}:F{40+row_iter}")
            ws.merge_cells(f"E{41+row_iter}:F{41+row_iter}")
            ws.merge_cells(f"E{42+row_iter}:F{42+row_iter}")
        ws.merge_cells(f"C{60+row_iter}:E{60+row_iter}")
        ws.merge_cells(f"C{61+row_iter}:E{61+row_iter}")
        ws.merge_cells(f"C{62+row_iter}:E{62+row_iter}")
        ws.merge_cells(f"J{4+row_iter}:K{4+row_iter}")

        set_size_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=5, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+23, max_row=row_iter+24, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=5, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+45, max_row=row_iter+50, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=2, max_col=6, size=helv_font9_bold)
        if bAmorphous:
            set_size_cells(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=3, max_col=3, size=helv_font9)
            set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=4, max_col=6, size=helv_font_bold)
            set_size_cells(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=5, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+60, max_row=row_iter+62, min_col=6, max_col=6, size=helv_font8)

        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=2, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=5, max_col=5, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=8, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+45, max_row=row_iter+56, min_col=3, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+60, max_row=row_iter+60, min_col=3, max_col=5, horizontal='left', vertical='center')

        ws.add_image(img3, f"J{59+row_iter}")
        ws.add_image(img_core, f"E{5+row_iter}")  
        sPrintArea = "A1:M272"
    else:
        # Ajustar anchos de columna
        column_widths = {
            'A': 0.5, 
            'B': 5.29, 
            'C': 14.35, 
            'D': 13.35, 
            'E': 13.35, 
            'F': 13.35, 
            'G': 13.35, 
            'H': 13.35, 
            'I': 13.35, 
            'J': 13.35,
            'K': 13.00, 
            'L': 1.5,
            'M': 2.5
        }
    
    
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.page_setup.scale = 78

        for row in range(1, 500):
            ws.row_dimensions[row].height = 13.5

        #QUOTE PAGE
        data = make_quote_page(device)

        for row in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=13):
            for cell in row:
                cell.fill = white_fill

        for row in ws.iter_rows(min_row=5, max_row=6, min_col=3, max_col=3):
            for cell in row:
                cell.fill = dot_fill

        for row in ws.iter_rows(min_row=9, max_row=9, min_col=9, max_col=9):
            for cell in row:
                cell.fill = dot_fill

        set_inside_border(ws, min_row=3, max_row=6, min_col=4, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=3, max_row=4, min_col=3, max_col=3, border=dotted_border)
        set_inside_border(ws, min_row=8, max_row=16, min_col=3, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=17, max_row=22, min_col=5, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=16, max_row=17, min_col=3, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=8, max_row=13, min_col=9, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=36, max_row=41, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=43, max_row=47, min_col=3, max_col=10, border=dotted_border)
        set_inside_border(ws, min_row=49, max_row=50, min_col=3, max_col=10, border=dotted_border)
        set_inside_border(ws, min_row=52, max_row=56, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=58, max_row=62, min_col=3, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=64, max_row=67, min_col=3, max_col=7, border=dotted_border)

        set_outer_border(ws, min_row=1, max_row=68, min_col=2, max_col=12, border=thin_border)

        set_outer_border(ws, min_row=3, max_row=6, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=4, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=5, max_row=6, min_col=3, max_col=3, border=thin_border)    
        set_outer_border(ws, min_row=3, max_row=6, min_col=4, max_col=5, border=thin_border)  
        set_outer_border(ws, min_row=3, max_row=6, min_col=6, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=6, min_col=8, max_col=9, border=thin_border)
        set_outer_border(ws, min_row=3, max_row=6, min_col=10, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=8, max_row=23, min_col=3, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=15, max_row=21, min_col=10, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=9, max_row=9, min_col=5, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=16, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=16, min_col=4, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=5, max_col=5, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=6, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=7, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=23, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=17, max_row=23, min_col=3, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=9, max_col=9, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=13, min_col=11, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=4, max_col=4, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=5, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=7, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=8, max_row=8, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=15, max_row=15, min_col=10, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=16, max_row=16, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=16, max_row=16, min_col=11, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=18, max_row=21, min_col=10, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=18, max_row=21, min_col=11, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=36, max_row=41, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=4, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=41, min_col=7, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=4, max_col=6, border=thin_border)
        set_outer_border(ws, min_row=36, max_row=36, min_col=7, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=43, max_row=47, min_col=3, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=47, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=47, min_col=4, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=43, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=43, max_row=43, min_col=4, max_col=10, border=thin_border)

        set_outer_border(ws, min_row=49, max_row=50, min_col=3, max_col=10, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=50, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=49, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=49, max_row=49, min_col=4, max_col=10, border=thin_border)
    
        set_outer_border(ws, min_row=52, max_row=56, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=56, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=56, min_col=4, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=52, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=52, max_row=52, min_col=4, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=58, max_row=62, min_col=3, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=62, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=62, min_col=4, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=58, min_col=3, max_col=3, border=thin_border)
        set_outer_border(ws, min_row=58, max_row=58, min_col=4, max_col=11, border=thin_border)

        set_outer_border(ws, min_row=64, max_row=67, min_col=3, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=64, max_row=64, min_col=3, max_col=7, border=thin_border)
        set_outer_border(ws, min_row=64, max_row=64, min_col=8, max_col=8, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=67, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=67, min_col=9, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=65, max_row=65, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=66, max_row=66, min_col=8, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=67, max_row=67, min_col=8, max_col=11, border=thin_border)

        # Llenar la hoja con datos y aplicar formato
        row_iter = 0
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_iter += 1
    
        ws.merge_cells("E8:F8")
        ws.merge_cells("I8:K8")
        ws.merge_cells("J15:K15")
        ws.merge_cells("J17:K17")
        ws.merge_cells("C18:D22")
        ws.merge_cells("D36:F36")
        ws.merge_cells("G36:K36")
        ws.merge_cells("D43:J43")
        ws.merge_cells("D52:K52")
        ws.merge_cells("D58:K58")
        ws.merge_cells("E64:G64")
        ws.merge_cells("E65:G65")
        ws.merge_cells("E66:G66")
        ws.merge_cells("E67:G67")
        ws.merge_cells("I65:K65")
        ws.merge_cells("I66:K66")
        ws.merge_cells("I67:K67")

        set_size_cells(ws, min_row=1, max_row=1, min_col=7, max_col=7, size=helv_font_bold_under)
        set_size_cells(ws, min_row=3, max_row=3, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=5, max_row=5, min_col=4, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=8, max_row=8, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=8, max_row=8, min_col=9, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=15, max_row=16, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=17, max_row=17, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=4, max_row=4, min_col=5, max_col=5, size=helv_font9)
        set_size_cells(ws, min_row=8, max_row=8, min_col=3, max_col=3, size=helv_font9_bold)
        set_size_cells(ws, min_row=37, max_row=37, min_col=6, max_col=6, size=helv_font8)
        set_size_cells(ws, min_row=43, max_row=43, min_col=3, max_col=3, size=helv_font9)
        set_size_cells(ws, min_row=50, max_row=50, min_col=3, max_col=3, size=helv_font9)

        d_values =[0.0,0.0]
        if device.losses.losses_warnings.warningValueCoreLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=10, max_row=10, min_col=11, max_col=11, size=helv_font_red)
        if device.losses.losses_warnings.warningValueLoadLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=11, max_row=11, min_col=11, max_col=11, size=helv_font_red)

        set_alignment_cells(ws, min_row=9, max_row=16, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=9, max_row=22, min_col=5, max_col=5, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=10, max_row=13, min_col=9, max_col=9, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=18, max_row=21, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=17, max_row=17, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=17, max_row=17, min_col=10, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=38, max_row=41, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=45, max_row=47, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=54, max_row=56, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=60, max_row=62, min_col=3, max_col=3, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=65, max_row=67, min_col=3, max_col=3, horizontal='left', vertical='center')

        row_iter += 2

        #WINDING SPECIFICATION
        sTapsLeadsTogether = "no empty" if pHCoil.fields["iBil"].get() > 200 and bHaveTaps else ""
        sTypeTaps =pWindH.changer.fields["sPerTaps"].get() if bHaveTaps else None

        data = make_winding_specification(device)
    
        for row in ws.iter_rows(min_row=row_iter+2, max_row=row_iter+72, min_col=2, max_col=11):
            for cell in row:
                cell.fill = dot_fill3

        # for row in ws.iter_rows(min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=4):
        #     for cell in row:
        #         cell.fill = white_fill
        set_inside_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=3, max_col=4, border=dotted_border)

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+72, min_col=2, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=6, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+12, max_row=row_iter+14, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+13, max_row=row_iter+14, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+13, max_row=row_iter+14, min_col=8, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+15, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+16, max_row=row_iter+16, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+17, max_row=row_iter+17, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+21, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+22, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+23, max_row=row_iter+24, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+15, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+16, max_row=row_iter+16, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+17, max_row=row_iter+17, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+21, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+22, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+23, max_row=row_iter+24, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+15, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+16, max_row=row_iter+16, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+17, max_row=row_iter+17, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+21, max_row=row_iter+21, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+22, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+23, max_row=row_iter+24, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+30, max_row=row_iter+30, min_col=3, max_col=4, border=thin_border, fillE = white_fill)  
    
        set_outer_border(ws, min_row=row_iter+31, max_row=row_iter+31, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+30, max_row=row_iter+30, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+31, max_row=row_iter+31, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=5, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+30, max_row=row_iter+30, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        #set_outer_border(ws, min_row=row_iter+31, max_row=row_iter+31, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        #set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=8, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+36, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+38, max_row=row_iter+38, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+40, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+40, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+40, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+41, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+41, min_col=5, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+38, max_row=row_iter+38, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+41, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+42, max_row=row_iter+42, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+43, max_row=row_iter+43, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+43, min_col=8, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+43, max_row=row_iter+43, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+45, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+45, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+45, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+46, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+46, max_row=row_iter+46, min_col=5, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+52, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+53, max_row=row_iter+53, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        if sTapsLeadsTogether!="":
            set_outer_border(ws, min_row=row_iter+53, max_row=row_iter+53, min_col=6, max_col=11, border=thin_border, fillE = yellow_fill)

        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+54, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+55, max_row=row_iter+56, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+59, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+65, max_row=row_iter+68, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=2, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+72, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+72, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+72, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        row_iter2 = 0
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row_iter2 += 1

    
        ws.merge_cells(f"F{2+row_iter}:H{2+row_iter}")
        ws.merge_cells(f"F{3+row_iter}:H{3+row_iter}")
        ws.merge_cells(f"F{4+row_iter}:H{4+row_iter}")
        ws.merge_cells(f"J{4+row_iter}:K{4+row_iter}")
        ws.merge_cells(f"C{12+row_iter}:D{12+row_iter}")
        ws.merge_cells(f"C{14+row_iter}:D{14+row_iter}")
        ws.merge_cells(f"E{12+row_iter}:G{12+row_iter}")
        ws.merge_cells(f"H{12+row_iter}:J{12+row_iter}")
        ws.merge_cells(f"H{23+row_iter}:I{23+row_iter}")
        ws.merge_cells(f"B{34+row_iter}:K{34+row_iter}")
        ws.merge_cells(f"H{38+row_iter}:I{38+row_iter}")
        ws.merge_cells(f"D{48+row_iter}:F{48+row_iter}")
        ws.merge_cells(f"C{49+row_iter}:D{49+row_iter}")
        ws.merge_cells(f"C{50+row_iter}:D{50+row_iter}")
        ws.merge_cells(f"E{49+row_iter}:F{49+row_iter}")
        ws.merge_cells(f"H{49+row_iter}:I{49+row_iter}")
        ws.merge_cells(f"H{50+row_iter}:I{50+row_iter}")
        ws.merge_cells(f"I{48+row_iter}:K{48+row_iter}")
        ws.merge_cells(f"J{49+row_iter}:K{49+row_iter}")
        ws.merge_cells(f"J{71+row_iter}:K{71+row_iter}")
        ws.merge_cells(f"J{72+row_iter}:K{72+row_iter}")    

        if sTapsLeadsTogether!="":
            ws.merge_cells(f"F{53+row_iter}:K{53+row_iter}")

        #set_size_cells(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=11, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+7, max_row=row_iter+7, min_col=8, max_col=8, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=6, max_col=6, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=9, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+31, max_row=row_iter+32, min_col=3, max_col=4, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+38, max_row=row_iter+43, min_col=8, max_col=9, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+53, max_row=row_iter+57, min_col=2, max_col=9, size=helv_font11)
        set_size_cells(ws, min_row=row_iter+53, max_row=row_iter+54, min_col=10, max_col=10, size=helv_font11)
        set_size_cells(ws, min_row=row_iter+56, max_row=row_iter+56, min_col=11, max_col=11, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=11, max_col=11, size=helv_font8)
        set_size_cells(ws, min_row=row_iter+62, max_row=row_iter+62, min_col=9, max_col=9, size=helv_font9)
        set_size_cells(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=10, max_col=11, size=helv_font9)
        if len(sTypeTaps)>16:
            set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=7, max_col=7, size=Cal_font8)

        if sTapsLeadsTogether!="":
            set_size_cells(ws, min_row=row_iter+53, max_row=row_iter+53, min_col=6, max_col=11, size=helv_font8)
        if ground_YHV==eLogical.TRUE:
            set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=8, max_col=8, size=helv_font9)    

        set_size_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=5, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=3, max_col=4, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=5, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=4, max_col=4, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=2, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+38, max_row=row_iter+39, min_col=8, max_col=9, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=8, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+65, max_row=row_iter+65, min_col=2, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=2, max_col=7, size=helv_font_bold)
        d_Values =[0.0,0.0]
        if  pWindH.winding_warnings.warningLeadsHVTaps(d_Values)==eWarnings.eWARNING:
            set_size_cells(ws, min_row=row_iter+60, max_row=row_iter+61, min_col=3, max_col=3, size=helv_font_red)

        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=6, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=3, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=10, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+12, max_row=row_iter+14, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+12, max_row=row_iter+12, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+15, max_row=row_iter+23, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+21, max_row=row_iter+22, min_col=11, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+23, max_row=row_iter+23, min_col=8, max_col=9, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+24, max_row=row_iter+24, min_col=10, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=3, max_col=3, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=5, max_col=5, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=8, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+27, max_row=row_iter+32, min_col=5, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+34, max_row=row_iter+36, min_col=2, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+39, max_row=row_iter+46, min_col=5, max_col=6, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+38, max_row=row_iter+43, min_col=8, max_col=9, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+41, max_row=row_iter+41, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+46, max_row=row_iter+46, min_col=3, max_col=4, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=4, max_col=6, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+48, max_row=row_iter+48, min_col=9, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+49, max_row=row_iter+50, min_col=3, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+52, max_row=row_iter+53, min_col=5, max_col=5, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=4, max_col=11, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=8, max_col=10, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=9, max_col=9, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+71, max_row=row_iter+71, min_col=8, max_col=8, horizontal='center', vertical='center')

        ws.add_image(img, f"J{69+row_iter}")

        row_iter += row_iter2+1
        row_iter += 2
        #ENGINEERING DATA

        data = make_engineering_data(device)

        for row in ws.iter_rows(min_row=row_iter+2, max_row=row_iter+72, min_col=2, max_col=11):
            for cell in row:
                cell.fill = white_fill

        set_inside_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+10, max_row=row_iter+12, min_col=4, max_col=6, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+10, max_row=row_iter+12, min_col=8, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=4, max_col=6, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=8, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=3, max_col=3, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=3, max_col=7, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+21, max_row=row_iter+23, min_col=4, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+26, max_row=row_iter+27, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+30, max_row=row_iter+31, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+33, max_row=row_iter+37, min_col=4, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+35, max_row=row_iter+43, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+40, max_row=row_iter+42, min_col=3, max_col=5, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=6, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+52, max_row=row_iter+57, min_col=4, max_col=9, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+61, max_row=row_iter+66, min_col=3, max_col=4, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=5, max_col=6, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=7, max_col=8, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+66, max_row=row_iter+67, min_col=10, max_col=11, border=dotted_border)
        set_inside_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=9, max_col=9, border=dotted_border)

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+72, min_col=2, max_col=11, border=thin_border)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+6, max_row=row_iter+8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+8, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+10, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+11, max_row=row_iter+11, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+10, max_row=row_iter+13, min_col=7, max_col=9, border=thick_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+15, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+15, max_row=row_iter+17, min_col=7, max_col=9, border=thick_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+20, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=4, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=4, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+21, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+21, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+21, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+21, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+20, max_row=row_iter+21, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+22, max_row=row_iter+23, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+19, max_row=row_iter+21, min_col=4, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+27, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+26, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+27, max_row=row_iter+27, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+25, max_row=row_iter+27, min_col=9, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+31, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+30, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+31, max_row=row_iter+31, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+31, min_col=9, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+34, min_col=4, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+35, min_col=4, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+37, min_col=4, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+37, min_col=3, max_col=3, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+43, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+43, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+35, max_row=row_iter+43, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+43, max_row=row_iter+43, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+42, min_col=3, max_col=5, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+47, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+44, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+46, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+46, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+47, max_row=row_iter+47, min_col=4, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+47, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+47, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+47, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+47, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+47, min_col=9, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+49, max_row=row_iter+49, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+49, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+57, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+51, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+57, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+57, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+58, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+58, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+58, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+58, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+51, min_col=3, max_col=3, border=thin_border, fillE = dot_fill)

        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+50, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+51, max_row=row_iter+52, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+53, max_row=row_iter+53, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+54, max_row=row_iter+55, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+56, max_row=row_iter+57, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+52, max_row=row_iter+57, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+57, min_col=4, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+57, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+57, min_col=8, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+60, min_col=3, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+66, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+67, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+67, min_col=7, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+60, max_row=row_iter+61, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+66, max_row=row_iter+67, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+66, max_row=row_iter+67, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+66, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=7, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=2, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+72, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+70, max_row=row_iter+72, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        # set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        # set_outer_border(ws, min_row=row_iter+71, max_row=row_iter+72, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        # set_outer_border(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=2, max_col=11, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        row_iter2 = 0
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_iter2 += 1
    
        ws.merge_cells(f"F{2+row_iter}:H{2+row_iter}")
        ws.merge_cells(f"F{3+row_iter}:H{3+row_iter}")
        ws.merge_cells(f"F{4+row_iter}:H{4+row_iter}")
        ws.merge_cells(f"G{11+row_iter}:G{13+row_iter}")
        ws.merge_cells(f"G{15+row_iter}:G{17+row_iter}")
        ws.merge_cells(f"D{19+row_iter}:J{19+row_iter}")
        ws.merge_cells(f"D{20+row_iter}:E{20+row_iter}")
        ws.merge_cells(f"D{33+row_iter}:H{33+row_iter}")
        ws.merge_cells(f"D{34+row_iter}:F{34+row_iter}")
        ws.merge_cells(f"G{34+row_iter}:H{34+row_iter}")
        ws.merge_cells(f"J{34+row_iter}:K{34+row_iter}")
        ws.merge_cells(f"C{44+row_iter}:E{44+row_iter}")
        ws.merge_cells(f"F{44+row_iter}:G{44+row_iter}")
        ws.merge_cells(f"J{44+row_iter}:K{44+row_iter}")
        ws.merge_cells(f"F{45+row_iter}:G{45+row_iter}")
        ws.merge_cells(f"F{46+row_iter}:G{46+row_iter}")
        ws.merge_cells(f"F{47+row_iter}:G{47+row_iter}")
        ws.merge_cells(f"J{45+row_iter}:K{45+row_iter}")
        ws.merge_cells(f"J{46+row_iter}:K{46+row_iter}")
        ws.merge_cells(f"D{47+row_iter}:E{47+row_iter}")
        ws.merge_cells(f"J{47+row_iter}:K{47+row_iter}")
        ws.merge_cells(f"C{49+row_iter}:K{49+row_iter}")
        ws.merge_cells(f"D{50+row_iter}:E{50+row_iter}")
        ws.merge_cells(f"F{50+row_iter}:G{50+row_iter}")
        ws.merge_cells(f"H{50+row_iter}:I{50+row_iter}")
        ws.merge_cells(f"J{50+row_iter}:K{50+row_iter}")
        ws.merge_cells(f"J{56+row_iter}:K{56+row_iter}")
        ws.merge_cells(f"J{57+row_iter}:K{57+row_iter}")
        ws.merge_cells(f"C{60+row_iter}:D{60+row_iter}")
        ws.merge_cells(f"E{60+row_iter}:F{60+row_iter}")
        ws.merge_cells(f"G{60+row_iter}:H{60+row_iter}")
        ws.merge_cells(f"J{59+row_iter}:K{59+row_iter}")
        ws.merge_cells(f"J{60+row_iter}:K{60+row_iter}")
        ws.merge_cells(f"J{61+row_iter}:K{61+row_iter}")
        ws.merge_cells(f"J{62+row_iter}:K{62+row_iter}")
        ws.merge_cells(f"C{69+row_iter}:F{69+row_iter}")
        ws.merge_cells(f"J{71+row_iter}:K{71+row_iter}")
        ws.merge_cells(f"J{72+row_iter}:K{72+row_iter}")

        set_size_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=3, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=8, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+14, max_row=row_iter+14, min_col=9, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+18, max_row=row_iter+18, min_col=3, max_col=9, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+19, max_row=row_iter+19, min_col=4, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+20, max_row=row_iter+20, min_col=4, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+25, max_row=row_iter+25, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=3, max_col=3, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+33, max_row=row_iter+33, min_col=4, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+34, max_row=row_iter+34, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+44, max_row=row_iter+44, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+49, max_row=row_iter+49, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=5, max_col=5, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+59, max_row=row_iter+59, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+60, max_row=row_iter+60, min_col=3, max_col=8, size=helv_font_bold)   
        set_size_cells(ws, min_row=row_iter+66, max_row=row_iter+67, min_col=10, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+69, max_row=row_iter+69, min_col=2, max_col=7, size=helv_font_bold)


        set_size_cells(ws, min_row=row_iter+44, max_row=row_iter+44, min_col=9, max_col=9, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+64, max_row=row_iter+64, min_col=9, max_col=9, size=helv_font8)
        set_size_cells(ws, min_row=row_iter+66, max_row=row_iter+67, min_col=9, max_col=9, size=helv_font8)

        if len(sTypeTaps)>16:
            set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=7, max_col=7, size=Cal_font8)

        if ground_YHV==eLogical.TRUE:
            set_size_cells(ws, min_row=row_iter+8, max_row=row_iter+8, min_col=8, max_col=8, size=helv_font9) 

        d_values = [0.0,0.0]
        if device.gradients.gradient_warnings.warningValueGradient(0,d_values) == eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+46, max_row=row_iter+46, min_col=4, max_col=4, size=helv_font_red)
        if device.gradients.gradient_warnings.warningValueGradient(1,d_values) == eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+47, max_row=row_iter+47, min_col=5, max_col=5, size=helv_font_red)
        if device.losses.losses_warnings.warningValueCoreLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+36, max_row=row_iter+36, min_col=4, max_col=4, size=helv_font_red)
        if device.losses.losses_warnings.warningValueLoadLosses(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+36, max_row=row_iter+36, min_col=5, max_col=5, size=helv_font_red)
        if device.core.core_warnings.warningFluxDensity(d_values)==eWarnings.eERROR:
            set_size_cells(ws, min_row=row_iter+22, max_row=row_iter+22, min_col=4, max_col=5, size=helv_font_red)

        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=2, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+35, max_row=row_iter+43, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+35, max_row=row_iter+43, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+51, max_row=row_iter+54, min_col=10, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+51, max_row=row_iter+54, min_col=11, max_col=11, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+66, min_col=3, max_col=3, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+66, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=5, max_col=5, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=6, max_col=6, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=7, max_col=7, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+61, max_row=row_iter+67, min_col=8, max_col=8, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+69, max_row=row_iter+70, min_col=8, max_col=8, horizontal='left', vertical='center')
        # Aplicar formato específico a celdas que necesitan diferente alineación
        # Celdas con texto largo - alineación izquierda

        ws.add_image(img2, f"J{69+row_iter}")
        
        row_iter += row_iter2+1

        #CORE DRAWING

        data = make_core_drawing(device)

        for row in ws.iter_rows(min_row=row_iter+1, max_row=row_iter+66, min_col=2, max_col=11):
            for cell in row:
                cell.fill = white_fill

        set_inside_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, border=dotted_border)

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+65, min_col=2, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+3, max_row=row_iter+3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+7, max_row=row_iter+22, min_col=4, max_col=9, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=6, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+29, max_row=row_iter+32, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+32, max_row=row_iter+32, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+36, min_col=6, max_col=6, border=thin_border, fillE = white_fill)  
        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+36, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+36, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+36, max_row=row_iter+36, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+33, max_row=row_iter+33, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+34, max_row=row_iter+37, min_col=9, max_col=11, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+39, max_row=row_iter+39, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=8, max_col=10, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+41, max_row=row_iter+42, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+44, max_row=row_iter+44, min_col=4, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+46, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+46, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+45, max_row=row_iter+46, min_col=6, max_col=6, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+50, max_row=row_iter+56, min_col=3, max_col=10, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+58, max_row=row_iter+60, min_col=3, max_col=8, border=thin_border, fillE = white_fill)

        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+62, min_col=2, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+65, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+65, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+63, max_row=row_iter+65, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+62, max_row=row_iter+63, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
        set_outer_border(ws, min_row=row_iter+64, max_row=row_iter+65, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

        # Llenar la hoja con datos y aplicar formato
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_iter+row_idx, column=col_idx, value=value)
                cell.font = helv_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells(f"F{2+row_iter}:H{2+row_iter}")
        ws.merge_cells(f"F{3+row_iter}:H{3+row_iter}")
        ws.merge_cells(f"F{4+row_iter}:H{4+row_iter}")
        ws.merge_cells(f"D{7+row_iter}:I{20+row_iter}")
        ws.merge_cells(f"E{22+row_iter}:H{22+row_iter}")
        ws.merge_cells(f"D{24+row_iter}:J{24+row_iter}")
        ws.merge_cells(f"F{26+row_iter}:G{26+row_iter}")
        ws.merge_cells(f"C{28+row_iter}:K{28+row_iter}")
        ws.merge_cells(f"I{33+row_iter}:K{33+row_iter}")
        ws.merge_cells(f"D{39+row_iter}:F{39+row_iter}")
        ws.merge_cells(f"H{39+row_iter}:J{39+row_iter}")
        ws.merge_cells(f"D{40+row_iter}:F{40+row_iter}")
        ws.merge_cells(f"H{40+row_iter}:J{40+row_iter}")
        ws.merge_cells(f"D{44+row_iter}:F{44+row_iter}")
        ws.merge_cells(f"C{62+row_iter}:F{62+row_iter}")

        set_size_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+4, max_row=row_iter+4, min_col=6, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=5, max_col=8, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+28, max_row=row_iter+28, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=3, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+33, max_row=row_iter+33, min_col=6, max_col=11, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=4, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+40, max_row=row_iter+40, min_col=8, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+44, max_row=row_iter+44, min_col=4, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+47, max_row=row_iter+48, min_col=6, max_col=6, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+49, max_row=row_iter+57, min_col=3, max_col=10, size=helv_font_bold)
        set_size_cells(ws, min_row=row_iter+62, max_row=row_iter+62, min_col=2, max_col=7, size=helv_font_bold)

        set_size_cells(ws, min_row=row_iter+29, max_row=row_iter+29, min_col=8, max_col=8, size=helv_font9_bold)
        set_size_cells(ws, min_row=row_iter+33, max_row=row_iter+33, min_col=8, max_col=8, size=helv_font9_bold)

        set_size_cells(ws, min_row=row_iter+24, max_row=row_iter+24, min_col=4, max_col=10, size=helv_font_red)

        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+3, min_col=2, max_col=2, horizontal='left', vertical='center')   
        set_alignment_cells(ws, min_row=row_iter+2, max_row=row_iter+2, min_col=4, max_col=4, horizontal='right', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+7, max_row=row_iter+21, min_col=3, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+26, max_row=row_iter+26, min_col=6, max_col=8, horizontal='center', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+34, max_row=row_iter+37, min_col=9, max_col=11, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+49, max_row=row_iter+56, min_col=3, max_col=3, horizontal='left', vertical='center')   
        set_alignment_cells(ws, min_row=row_iter+58, max_row=row_iter+58, min_col=4, max_col=4, horizontal='left', vertical='center') 
        set_alignment_cells(ws, min_row=row_iter+60, max_row=row_iter+60, min_col=3, max_col=3, horizontal='left', vertical='center') 
        set_alignment_cells(ws, min_row=row_iter+64, max_row=row_iter+64, min_col=9, max_col=10, horizontal='left', vertical='center')
        set_alignment_cells(ws, min_row=row_iter+62, max_row=row_iter+62, min_col=3, max_col=3, horizontal='left', vertical='center')

        for row in range(1, 66):
            ws.row_dimensions[row].height = 14.25  

        ws.add_image(img3, f"J{62+row_iter}")
        ws.add_image(img_core, f"D{8+row_iter}")  
        sPrintArea = "A1:L292"

    # Guardar el archivo
    filename = f"{sAddress}\\{device.fields['so_number'].get()}_reports.xlsx"
    filenamePDF = f"{sAddress}\\{device.fields['so_number'].get()}_reports.pdf"
    if wb is not None:
        wb.save(filename)

    # Intenta usar una instancia ya abierta
    try:
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        excel = win32.Dispatch("Excel.Application")

    excel.Visible = True
    excel.DisplayAlerts = False

    # Verificar si el archivo ya está abierto
    workbook = None
    for wb_open in excel.Workbooks:
        if os.path.abspath(wb_open.FullName).lower() == os.path.abspath(filename).lower():
            workbook = wb_open
            break

    # Si no está abierto, abrirlo
    if workbook is None:
        workbook = excel.Workbooks.Open(filename)

    ws = workbook.Sheets("Sheet1")
    ws.PageSetup.PrintArea = sPrintArea

    # Exportar a PDF
    ws.ExportAsFixedFormat(0, filenamePDF)  # 0 = PDF

    if os.path.exists(filenamePDF):
        os.startfile(filenamePDF)

    if not is_folder_open(sAddress):
        os.startfile(sAddress)

    return filename

