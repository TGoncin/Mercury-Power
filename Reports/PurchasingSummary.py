import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
from datetime import datetime
import win32com.client as win32
import os
from Resources.resources import resource_path
from Definitions.enums import eConnectionType,eArrayType,eLogical,eBIL,eWarnings,eTransformerFamily,eWindingSelection,eTapExtremes,eSequence,eChangerCapacity,eMaterial,eLaminationType,eTypeConductor,eCoating,eDesignSpec,eCoolingType,eSingleCoolingType,eTemperatureRise
from Definitions.functions import address,ADDRESSZ,convertCMInch,is_folder_open
from datetime import datetime
import math

def make_purchasing_summary(device = None):
    def set_inside_border(ws,min_row, max_row, min_col, max_col, border = None, fillE = None):
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border  # apply same border to every cell
                if fillE is not None:
                    cell.fill = fillE


    def set_outer_border(ws, min_row, max_row, min_col, max_col, border, fillE = None):
        """
        Applies a border only around the perimeter of a cell range.

        ws : Worksheet
        min_row : starting row of the range
        max_row : ending row of the range
        min_col : starting column of the range
        max_col : ending column of the range
        border : Border object (e.g., Border(left=Side(...)))
        """
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                # Borde exterior superior
                top = border.top if row == min_row else cell.border.top
                # Borde exterior inferior
                bottom = border.bottom if row == max_row else cell.border.bottom
                # Borde exterior izquierdo
                left = border.left if col == min_col else cell.border.left
                # Borde exterior derecho
                right = border.right if col == max_col else cell.border.right
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
                if fillE is not None:
                    cell.fill = fillE

    def set_size_cells(ws, min_row, max_row, min_col, max_col, size):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = size

    def set_alignment_cells(ws, min_row, max_row, min_col, max_col, horizontal, vertical):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    sAddress = ADDRESSZ#address(device.fields['so_number'].get())
    so_number = str(device.fields['so_number'].get())  # ensure it's a string
    sAddress = os.path.join(sAddress, so_number)
    os.makedirs(sAddress, exist_ok=True)  

    #Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    img_path = resource_path("Resources/logo_reports.png")
    img_pil = PILImage.open(img_path)

    dpi = 96
    width_inch = 1.48
    height_inch = 0.37
    width_px = int(width_inch * dpi)
    height_px = int(height_inch * dpi)

    img_pil = img_pil.resize((width_px, height_px), PILImage.Resampling.LANCZOS)

    img_bytes = BytesIO()
    img_pil.save(img_bytes, format="PNG")
    img_bytes.seek(0)

    img = Image(img_bytes)

    ws.page_margins = PageMargins(
                        left=0.25,
                        right=0.25,
                        top=0.25,
                        bottom=0.25,
                        header=0,
                        footer=0
                    )

    # Definir estilos
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))


    thick_border = Border(left=Side(style='medium'),   # también puedes usar 'thick'
                          right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    dotted_border = Border(
                            left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

    # Fondo de puntitos (pattern fill)
    dot_fill = PatternFill(
                            fill_type="gray125",  # patrón de puntos
                            start_color="000000", # color de los puntos: negro
                            end_color="FFFFFF"    # color de fondo: blanco (o transparente)
                        )
    
    white_fill = PatternFill(
                            fill_type="solid",
                            start_color="FFFFFF",
                            end_color="FFFFFF"
                        )

    helv_font11 = Font(name="Helvetica", size=11)
    helv_font = Font(name="Helvetica", size=10)
    helv_font_bold = Font(name="Helvetica", size=10, bold = True)
    helv_font_red = Font(name="Helvetica", size=10,color="FF0000")
    helv_font9 = Font(name="Helvetica", size=9)
    helv_font9_bold = Font(name="Helvetica", size=9, bold = True)
    helv_font8 = Font(name="Helvetica", size=8)
    helv_font8_bold = Font(name="Helvetica", size=8, bold = True)

    # Ajustar anchos de columna
    column_widths = {
        'A': 2.57, 
        'B': 6, 
        'C': 12.29, 
        'D': 10.14, 
        'E': 12.29, 
        'F': 12.29, 
        'G': 12.29, 
        'H': 12.29, 
        'I': 12.29, 
        'J': 14.86,
        'K': 12.5, 
        'L': 0.92
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.print_area = "A1:L63"
    ws.page_setup.scale = 83  # 70%
    # Opcional: orientación y ajuste a ancho/alto
    #ws.page_setup.orientation = 'portrait'  # 'landscape' también
    # Datos de la hoja (A1:L69)

    pWindH = device.getHVWind()
    pWindL = device.getLVWind()
    pLCoil = pWindL.coils[0]
    pHCoil = pWindH.coils[0]
    pLCond = pLCoil.conductor[0]
    pHCond = pHCoil.conductor[0]

    tCoolingTypeE = device.cooling.fields['t_CoolingType'].get()
    if tCoolingTypeE[0]==eSingleCoolingType.COOL_ONAN:
        sCooling = "ONAN"
    elif tCoolingTypeE[0]==eSingleCoolingType.COOL_KNAN:
        sCooling = "KNAN"
    iSteps = device.cooling.fields['iCoolingSteps'].get()
    if iSteps>1:
        if tCoolingTypeE[1]==eSingleCoolingType.COOL_ONAF:
            sCooling += " ONAF"
        elif tCoolingTypeE[1]==eSingleCoolingType.COOL_KNAF:
            sCooling += " KNAF"

    if iSteps>2:
        if tCoolingTypeE[2]==eSingleCoolingType.COOL_ONAF:
            sCooling += " ONAF"
        elif tCoolingTypeE[2]==eSingleCoolingType.COOL_KNAF:
            sCooling += " KNAF"

    sCooling += f" {device.fields['frequency'].get()} Hz "
    sCooling += f" {device.cooling.fields['tTemperatureRise'].get()[0].value} ° C"

    sCustomSpec = f"Cust Spec:   {device.guarantee.fields['sCustomerSpec'].get()}"

    sHVoltage1 = pWindH.fields['sVoltage1'].get()#2 Bushing" if pWindH.fields['iBil_Under'].get()==pWindH.fields['iBIL'].get().value else "Single Bushing"
    sLVoltage1 = pWindL.fields['sVoltage1'].get()#f'{pWindL.fields['dVoltage'].get()} /' if pWindL.fields['iNumBushings'].get() > 2 else None
    sLVoltage2 = pWindL.fields['sVoltage2'].get()#pWindL.fields['dVoltage'].get() / 2 if pWindL.fields['iNumBushings'].get() > 2 else pWindL.fields['dVoltage'].get()

    sRev = ['']*3
    sRevDescr = ['']*3
    sDateDescr = ['']*3
    iRev = device.fields['i_RevisionWS'].get()[0]
    sRev[0] = 1 if iRev!=0 else ""
    sRev[1] = 2 if iRev>=2 else ""
    sRev[2] = 3 if iRev==3 else ""
    sRevDescr[0] = device.fields['s_DataDescr0'].get()[0] if iRev!=0 else ""
    sRevDescr[1] = device.fields['s_DataDescr0'].get()[1] if iRev>=2 else ""
    sRevDescr[2] = device.fields['s_DataDescr0'].get()[2] if iRev==3 else ""
    sDateDescr[0] = device.fields['s_DateDescr0'].get()[0] if iRev!=0 else ""
    sDateDescr[1] = device.fields['s_DateDescr0'].get()[1] if iRev>=2 else ""
    sDateDescr[2] = device.fields['s_DateDescr0'].get()[2] if iRev==3 else ""
    sDP = f"{device.fields['dLatestVersionNo'].get():.2f}" if device.fields['sDPSpecial'].get() != "M" else f"{device.fields['dLatestVersionNo'].get():.2f} M"
    sDwg = f"PS-{device.fields['so_number'].get()}"
    date = datetime.strptime(device.fields['original_issue_date'].get(), "%m/%d/%Y")
    sformat_date = date.strftime("%b %#d/%y")
    bTapsFCBN = True if pWindH.have_taps() and pWindH.changer.fields['tCapacity'].get == eChangerCapacity.FCBN else False
    dHVLineAmps =  f"{pWindH.fields['pdLineAmp'].get()[pWindH.getNomPos()]:.2f}"
    dHVPhAmps = pWindH.fields['sLineAmp'].get()
    dHVLineAmpsFCBN =  f"{pWindH.fields['pdLineAmp'].get()[0]:.2f}" if bTapsFCBN else ""
    dHVPhAmpsFCBN = f"{pWindH.fields['pdPhaseAmp'].get()[0]:.2f}" if bTapsFCBN else ""
    dLVLineAmps =  f"{pWindL.fields['pdLineAmp'].get()[pWindL.getNomPos()]:.2f}"
    dLVPhAmps = pWindL.fields['sLineAmp'].get()

    def calc_kva_af(Cooling_Type, P131, pWind, CSA_Spec,iPhases)-> tuple[float, float,float]:
        kVA_Nom = pWind.fields['d_kVA'].get()[0]
        dVolt = pWind.fields['dVoltage'].get()
        dkVA = 0.0
        dPhAmps = 0.0
        dLineAmps = 0.0
        if "AF" in Cooling_Type:
            if P131 != 0:
                dkVA = round((1 + P131 / 100) * kVA_Nom, 2)
            else:
                if CSA_Spec == "ANSI":
                    if kVA_Nom <= 2000:
                        dkVA = round(kVA_Nom * 1.15, 0)
                    elif kVA_Nom <= 10000:
                        dkVA = round(kVA_Nom * 1.25, 1)
                    else:
                        dkVA = 0.0
                else:
                    dkVA = (round(kVA_Nom / 100 * 4 / 3, 1) * 100)
            dPhAmps = dkVA/iPhases/dVolt*1000
            dLineAmps = dPhAmps*(math.sqrt(3) if pWind.fields['type_connection'].get()==eConnectionType.DELTA else 1)
        return dkVA,dPhAmps,dLineAmps

    def calc_kva_af_slash(Cooling_Type, pWind,CSA_Spec, iPhases)-> tuple[float, float,float]:
        kVA_Nom = pWind.fields['d_kVA'].get()[0]
        dVolt = pWind.fields['dVoltage'].get()
        dkVA = 0.0
        dPhAmps = 0.0
        dLineAmps = 0.0
        if "AF_" in Cooling_Type:
            if CSA_Spec == "ANSI":
                if kVA_Nom <= 2000:
                    dkVA = round(kVA_Nom * 1.3, 0)
                elif kVA_Nom <= 10000:
                    dkVA = round(kVA_Nom * 1.5, 1)
                else:
                    dkVA = None  # igual que Excel: no hay caso definido
            else:
                dkVA = round(kVA_Nom * (5 / 3), 1)
            dPhAmps = dkVA/iPhases/dVolt*1000
            dLineAmps = dPhAmps*(math.sqrt(3) if pWind.fields['type_connection'].get()==eConnectionType.DELTA else 1)

        return dkVA,dPhAmps,dLineAmps

    def calc_kva_af_fcbn(bTapsFCBN,Cooling_Type, kVA,pWind, iPhases)-> tuple[float, float]:
        dPhAmps = 0.0
        dLineAmps = 0.0
        if bTapsFCBN==False:
            return dPhAmps,dLineAmps

        if "AF" in Cooling_Type:
            dPhAmps = dkVA/iPhases/pWind.fields['d_LinekVPerStep'].get()*1000
            dLineAmps = dPhAmps*(math.sqrt(3) if pWind.fields['type_connection'].get()==eConnectionType.DELTA else 1)
        else:
            dkVA = 0.0
            dPhAmps = 0.0
            dLineAmps = 0.0
        return dPhAmps,dLineAmps

    dkVA1,dPhAmpsHV1,dLineAmpsHV1 = calc_kva_af(device.cooling.fields['tCoolingTypes'].get().name,0.0,pWindH,device.guarantee.fields['tDesignSpec'].get().name,device.fields['iNoPhases'].get())
    skVA1 = f"{dkVA1:.2f}" if dkVA1!=0.0 else ""
    sPhAmpsHV1 = f"{dPhAmpsHV1:.2f}" if dPhAmpsHV1!=0.0 else ""
    sLineAmpsHV1 = f"{dLineAmpsHV1:.2f}" if dLineAmpsHV1!=0.0 else ""

    dkVA1,dPhAmpsLV1,dLineAmpsLV1 = calc_kva_af(device.cooling.fields['tCoolingTypes'].get().name,0.0,pWindL,device.guarantee.fields['tDesignSpec'].get().name,device.fields['iNoPhases'].get())
    skVA1 = f"{dkVA1:.2f}" if dkVA1!=0.0 else ""
    sPhAmpsLV1 = f"{dPhAmpsLV1:.2f}" if dPhAmpsLV1!=0.0 else ""
    sLineAmpsLV1 = f"{dLineAmpsLV1:.2f}" if dLineAmpsLV1!=0.0 else ""

    dPhAmpsHVFCBN1,dLineAmpsHVFCBN1 = calc_kva_af_fcbn(bTapsFCBN,device.cooling.fields['tCoolingTypes'].get().name,dkVA1,pWindH,device.fields['iNoPhases'].get())
    sPhAmpsHVFCBN1 = f"{dPhAmpsHVFCBN1:.2f}" if dPhAmpsHVFCBN1!=0.0 else ""
    sLineAmpsHVFCBN1 = f"{dLineAmpsHVFCBN1:.2f}" if dLineAmpsHVFCBN1!=0.0 else ""

    dkVA2,dPhAmpsHV2,dLineAmpsHV2 = calc_kva_af_slash(device.cooling.fields['tCoolingTypes'].get().name,pWindH,device.guarantee.fields['tDesignSpec'].get().name,device.fields['iNoPhases'].get())
    skVA2 = f"{dkVA2:.2f}" if dkVA2!=0.0 else ""
    sPhAmpsHV2 = f"{dPhAmpsHV2:.2f}" if dPhAmpsHV2!=0.0 else ""
    sLineAmpsHV2 = f"{dLineAmpsHV2:.2f}" if dLineAmpsHV2!=0.0 else ""

    dkVA2,dPhAmpsLV2,dLineAmpsLV2 = calc_kva_af_slash(device.cooling.fields['tCoolingTypes'].get().name,pWindL,device.guarantee.fields['tDesignSpec'].get().name,device.fields['iNoPhases'].get())
    skVA2 = f"{dkVA2:.2f}" if dkVA2!=0.0 else ""
    sPhAmpsLV2 = f"{dPhAmpsLV2:.2f}" if dPhAmpsLV2!=0.0 else ""
    sLineAmpsLV2 = f"{dLineAmpsLV2:.2f}" if dLineAmpsLV2!=0.0 else ""

    dPhAmpsHVFCBN2,dLineAmpsHVFCBN2 = calc_kva_af_fcbn(bTapsFCBN,device.cooling.fields['tCoolingTypes'].get().name,dkVA2,pWindH,device.fields['iNoPhases'].get())
    sPhAmpsHVFCBN2 = f"{dPhAmpsHVFCBN2:.2f}" if dPhAmpsHVFCBN2!=0.0 else ""
    sLineAmpsHVFCBN2 = f"{dLineAmpsHVFCBN2:.2f}" if dLineAmpsHVFCBN2!=0.0 else ""

    s_WeightsE = ['']*6
    s_WeightsE[0] = f"{device.fields['dWeightKr0022'].get():.1f}" if device.fields['dWeightKr0022'].get()>0.0 else ""
    s_WeightsE[1] = f"{device.fields['dWeightKr005'].get():.1f}" if device.fields['dWeightKr005'].get()>0.0 else ""
    s_WeightsE[2] = f"{device.fields['dWeightKr010'].get():.1f}" if device.fields['dWeightKr010'].get()>0.0 else ""
    s_WeightsE[3] = f"{device.fields['dWeightKr015'].get():.1f}" if device.fields['dWeightKr015'].get()>0.0 else ""
    s_WeightsE[4] = f"{device.core.fields['dWeightsPB1_8'].get():.1f}" if device.core.fields['dWeightsPB1_8'].get()>0.0 else ""
    s_WeightsE[5] = f"{device.core.fields['dWeightsPB1_16'].get():.1f}" if device.core.fields['dWeightsPB1_16'].get()>0.0 else ""

    s_DuctMaterial = ['']*2
    s_DuctMaterialfu = ['']*2
    dWidthDucts = convertCMInch(pLCoil.fields['dWidthDucts'].get()/10,1)
    s_DuctMaterial[0] = f"{dWidthDucts}'' x 52'' Wide Rolls" if dWidthDucts>0.0 else ""
    s_DuctMaterial[1] = f"{pHCoil.fields['dUOHVDuct'].get()}'' x 52'' Wide Rolls" if pHCoil.fields['dUOHVDuct'].get()>0.0 else ""
    s_DuctMaterialfu[0] = f"{device.core.fields['d_DuctsDM'].get()[0]:.1f}" if device.core.fields['d_DuctsDM'].get()[0]>0.0 else ""
    s_DuctMaterialfu[1] = f"{device.core.fields['d_DuctsDM'].get()[1]:.1f}" if device.core.fields['d_DuctsDM'].get()[1]>0.0 else ""
    sMetric = f"{round(pLCond.fields['dThickStrand'].get() * 25.4, 2)}  x  {round(pLCond.fields['dWidthStrand'].get() * 25.4, 2)} mm" if pLCond.fields['iType'].get()==eTypeConductor.STRIP else ""
    tTypeCondLV = (
        "LV "
        + ("Copper" if pLCond.fields['tMatAluminum'].get() == eMaterial.CU else "Aluminum")
        + " "
        + ("Strip" if pLCond.fields['iType'].get() == eTypeConductor.STRIP else "Wire")
    )
    tTypeCondHV = (
        "HV "
        + ("Copper" if pHCond.fields['tMatAluminum'].get() == eMaterial.CU else "Aluminum")
        + " "
        + ("Strip" if pHCond.fields['iType'].get() == eTypeConductor.STRIP else "Wire")
    )
    sMetricCL = f"{round(pWindL.fields['dThicknessLeads'].get() * 25.4, 1)}  x  {round(pWindL.fields['dWidthLeads'].get() * 25.4, 1)} mm"
    sInchesCL = f"{pWindL.fields['dThicknessLeads'].get()}  x  {pWindL.fields['dWidthLeads'].get()}''"
    skgUnitCL = f"{pWindL.fields['dWeight_LBLead'].get():.1f}" if pWindL.fields['dWeight_LBLead'].get()>0.0 else ""
    slbUnitCL = f"{(pWindL.fields['dWeight_LBLead'].get()*2.2046):.1f}" if pWindL.fields['dWeight_LBLead'].get()>0.0 else ""

    def split_string(text):
        max_len = 91
        words = text.split()
        parts = []
        temp = ""

        for w in words:
            # Calcular longitud si añadimos la palabra
            if temp:
                new_len = len(temp) + 1 + len(w)
            else:
                new_len = len(w)

            if new_len > max_len:
                # Guardar parte y empezar nueva
                parts.append(temp)
                temp = w
            else:
                temp += (" " if temp else "") + w

            # Si ya tenemos 3 partes, agregamos todo lo que quede a la última
            if len(parts) == 2:
                remaining_words = words[words.index(w)+1:]
                if remaining_words:
                    temp += " " + " ".join(remaining_words)
                break

        if temp:
            parts.append(temp)

        # Asegurar que siempre sean 3 elementos
        while len(parts) < 3:
            parts.append("")

        return parts

    sPSComments = device.fields['sPurchasingComments'].get()
    sComments4PS = split_string(sPSComments)

    data = [
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, "For:", None,  device.fields['customer_name'].get() , None, "Purchasing Summary", None, None, None, "SO #:", device.fields['so_number'].get(),None],
        [None, "Code:", None, device.fields['customer_code'].get(), None, "Standard Shell Type, Layer Winding", None, None, None, "Primary Spec:", device.guarantee.fields['tDesignSpec'].get().value,None],
        [None, None, None, None, None, sCooling, None, None, None, sCustomSpec,None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, "Taps", "HV", "LV", None,None,None],
        [None, None, "SO #", "kVA", "Phases", "Type", pWindH.changer.get_Type() if pWindH.have_taps() else "None", sHVoltage1, sLVoltage1,"HV-BIL", "LV-BIL",None],
        [None, None, device.fields['so_number'].get(), pWindH.fields['d_kVA'].get()[0], device.fields['iNoPhases'].get(), device.fields['transformer_family'].get().value,pWindH.changer.fields['tCapacity'].get().value if pWindH.have_taps() else None, pWindH.fields['dVoltage'].get(), sLVoltage2,  pWindH.fields['iBIL'].get().value, pWindL.fields['iBIL'].get().value, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Rise", "kVA","HV Line Amps","HV Ph Amps", None, "HV LineAmps", "HV Ph Amps", "LV Line Amps", "LV PH Amps", None, None],
        [None, None, f"{device.cooling.fields['tTemperatureRises'].get().value}°C", pWindH.fields['d_kVA'].get()[0], dHVLineAmps,dHVPhAmps , "FCBN", dHVLineAmpsFCBN, dHVPhAmpsFCBN, dLVLineAmps, dLVPhAmps, None],
        [None, None, None, skVA1, sLineAmpsHV1, sPhAmpsHV1, None, sPhAmpsHVFCBN1, sLineAmpsHVFCBN1, sLineAmpsLV1, sPhAmpsLV1, None],
        [None, None, None, skVA2, sLineAmpsHV2, sPhAmpsHV2, None, sPhAmpsHVFCBN2, sLineAmpsHVFCBN2, sLineAmpsLV2, sPhAmpsLV2, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, "FCBN", None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "HV Current Ratings", None, "LV Current Ratings", None, None, None, None, None, None, None],
        [None, None, "Line Amps", "Ph Amps", "Line Amps", "Ph Amps", None, None, None, None, None, None],
        [None, None, dHVLineAmps, dHVPhAmps, dLVLineAmps, dLVPhAmps, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Insulating Materials", None, None, None, None, "Duct Material", None, "ft / Unit", None, None],
        [None, None, "Description", None, "Lb / Unit", None, None, s_DuctMaterial[0],None, s_DuctMaterialfu[0], None, None],
        [None, None, "Kraft DD Epoxy - .0022''", None, s_WeightsE[0], None, None, s_DuctMaterial[1],None, s_DuctMaterialfu[1], "" if s_DuctMaterial[1]=="" else "A-duct", None],
        [None, None, "Kraft DD Epoxy - .005''", None, s_WeightsE[1], None, None, None, None, None, None, None],
        [None, None, "Kraft DD Epoxy - .010''", None, s_WeightsE[2], None, None, None, None, None, None, None],
        [None, None, "Kraft DD Epoxy - .015''", None, s_WeightsE[3], None, None, None, None, None, None, None],
        [None, None, "Pressboard - 1/8''", None, s_WeightsE[4], None, None, "Pressboard Weights Include C & C Pkg's", None, None, None, None],
        [None, None, "Pressboard - 1/16''", None, s_WeightsE[5], None, None, "Bottom Clamp Pkg and Frame Insulation", None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Conductor", None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Metric", None, tTypeCondLV, None, "kg/Unit", "Lb/Unit", "High", "Wide", None, None],
        [None, None, None, sMetric, pLCond.fields['sDescription'].get(), None, f"{(pLCond.fields['dWeight'].get()/2.2046):.0f}", f"{pLCond.fields['dWeight'].get():.0f}", pLCoil.fields['iHigh'].get(), pLCoil.fields['iWide'].get(), None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, tTypeCondHV, None, "kg/Unit", "Lb/Unit", "High", "Wide", None, None],
        [None, None, None, None, pHCond.fields['sDescription'].get(), None, f"{(pHCond.fields['dWeight'].get()/2.2046):.0f}", f"{pHCond.fields['dWeight'].get():.0f}", pHCoil.fields['iHigh'].get(), pHCoil.fields['iWide'].get(), None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Leads & Bus Material", None, None, None, None, None, None, None, None, None],
        [None, None, "Material", "Metric", None, "Inches", None, "Kg / Unit", "Lb / unit", None, None, None],
        [None, None, pWindL.fields['sCoilLdMat'].get(), sMetricCL, None, sInchesCL, None, skgUnitCL, slbUnitCL, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "Comments From Engineering:", None, None, None, None, None, None, None, None, None],
        [None, None, sComments4PS[0], None, None, None, None, None, None, None, None, None],
        [None, None, sComments4PS[1], None, None, None, None, None, None, None, None, None],
        [None, None, sComments4PS[2], None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [None, "Rev", "Revision Description", None, None, "Rev By/Date", "Design By:", device.fields['issued_by'].get(), None, None, None, None],
        [None, sRev[0], sRevDescr[0], None, None, sDateDescr[0], "Design Date:", sformat_date, None, None, None, None],
        [None, sRev[1], sRevDescr[1], None, None, sDateDescr[1], "DP Ver. No.", None, "Title: Purchasing Summary",None ,"Revision No", None],
        [None, sRev[2], sRevDescr[2], None, None, sDateDescr[2], sDP, None, "Drawing No:", sDwg, iRev, None]
    ]
    
    for row in ws.iter_rows(min_row=2, max_row=62, min_col=2, max_col=11):
        for cell in row:
            cell.fill = white_fill

    set_inside_border(ws, min_row=2, max_row=2, min_col=6, max_col=8, border=dotted_border)
    set_inside_border(ws, min_row=10, max_row=12, min_col=4, max_col=6, border=dotted_border)
    set_inside_border(ws, min_row=10, max_row=12, min_col=8, max_col=11, border=dotted_border)
    set_inside_border(ws, min_row=16, max_row=17, min_col=4, max_col=6, border=dotted_border)
    set_inside_border(ws, min_row=16, max_row=17, min_col=8, max_col=11, border=dotted_border)
    set_inside_border(ws, min_row=20, max_row=21, min_col=3, max_col=6, border=dotted_border)
    set_inside_border(ws, min_row=25, max_row=30, min_col=3, max_col=5, border=dotted_border)
    set_inside_border(ws, min_row=36, max_row=37, min_col=3, max_col=8, border=dotted_border)
    set_inside_border(ws, min_row=40, max_row=41, min_col=3, max_col=8, border=dotted_border)

    set_outer_border(ws, min_row=2, max_row=62, min_col=2, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=2, max_row=2, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=2, max_row=3, min_col=2, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=2, max_row=4, min_col=6, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=2, max_row=2, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=3, max_row=3, min_col=10, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=4, max_row=4, min_col=10, max_col=11, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=7, max_row=8, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=7, max_row=8, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=7, max_row=8, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=7, max_row=8, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=6, max_row=8, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=6, max_row=8, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=6, max_row=8, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=7, max_row=8, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=7, max_row=8, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=10, max_row=10, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=11, max_row=11, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=11, max_row=13, min_col=4, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=10, max_row=13, min_col=7, max_col=9, border=thick_border, fillE = white_fill)

    set_outer_border(ws, min_row=15, max_row=15, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=4, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=11, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=15, max_row=17, min_col=7, max_col=9, border=thick_border, fillE = white_fill)

    set_outer_border(ws, min_row=19, max_row=21, min_col=3, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=19, max_row=19, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=19, max_row=19, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=20, max_row=21, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=20, max_row=21, min_col=5, max_col=6, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=23, max_row=23, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=24, max_row=24, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=24, max_row=24, min_col=5, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=25, max_row=30, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=25, max_row=30, min_col=5, max_col=5, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=23, max_row=23, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=23, max_row=23, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=24, max_row=24, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=24, max_row=24, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=25, max_row=25, min_col=8, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=25, max_row=25, min_col=10, max_col=10, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=29, max_row=30, min_col=8, max_col=11, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=34, max_row=42, min_col=3, max_col=11, border=thin_border, fillE = dot_fill)
    set_outer_border(ws, min_row=33, max_row=33, min_col=3, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=37, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=37, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=37, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=37, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=36, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=36, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=35, max_row=35, min_col=3, max_col=10, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=40, max_row=41, min_col=3, max_col=4, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=41, min_col=5, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=41, min_col=7, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=41, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=40, min_col=9, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=40, min_col=10, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=39, max_row=39, min_col=5, max_col=10, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=46, max_row=46, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=47, min_col=3, max_col=9, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=48, min_col=3, max_col=3, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=48, min_col=4, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=48, min_col=6, max_col=7, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=48, min_col=8, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=47, max_row=48, min_col=9, max_col=9, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=54, max_row=57, min_col=3, max_col=11, border=thin_border, fillE = white_fill)

    set_outer_border(ws, min_row=59, max_row=59, min_col=2, max_col=2, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=59, max_row=59, min_col=3, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=59, max_row=59, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=59, max_row=60, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=59, max_row=60, min_col=9, max_col=11, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=60, max_row=62, min_col=2, max_col=5, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=60, max_row=62, min_col=6, max_col=6, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=61, max_row=62, min_col=7, max_col=8, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=61, max_row=62, min_col=9, max_col=10, border=thin_border, fillE = white_fill)
    set_outer_border(ws, min_row=61, max_row=62, min_col=11, max_col=11, border=thin_border, fillE = white_fill)

    # Llenar la hoja con datos y aplicar formato
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = helv_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells("C2:D2")
    ws.merge_cells("F2:H2")
    ws.merge_cells("F3:H3")
    ws.merge_cells("F4:H4")
    ws.merge_cells("G11:G13")   
    ws.merge_cells("G15:G17")
    ws.merge_cells("C19:D19")
    ws.merge_cells("E19:F19")
    ws.merge_cells("H23:I23")
    ws.merge_cells("H24:I24")
    ws.merge_cells("H25:I25")
    ws.merge_cells("C23:E23")
    ws.merge_cells("C24:D24")
    ws.merge_cells("C25:D25")
    ws.merge_cells("C26:D26")
    ws.merge_cells("C27:D27")
    ws.merge_cells("C28:D28")
    ws.merge_cells("C29:D29")
    ws.merge_cells("C30:D30")
    ws.merge_cells("H29:K29")
    ws.merge_cells("H30:K30")
    ws.merge_cells("C33:K33")
    ws.merge_cells("C35:D35")
    ws.merge_cells("C36:D36")
    ws.merge_cells("C37:D37")
    ws.merge_cells("E35:F35")
    ws.merge_cells("E36:F36")
    ws.merge_cells("E37:F37")
    ws.merge_cells("C40:D40")
    ws.merge_cells("C41:D41")
    ws.merge_cells("E39:F39")
    ws.merge_cells("E40:F40")
    ws.merge_cells("E41:F41")
    ws.merge_cells("C46:I46")
    ws.merge_cells("D47:E47")
    ws.merge_cells("D48:E48")
    ws.merge_cells("F47:G47")
    ws.merge_cells("F48:G48")
    ws.merge_cells("C54:K54")
    ws.merge_cells("C55:K55")
    ws.merge_cells("C56:K56")
    ws.merge_cells("C57:K57")
    ws.merge_cells("C58:K58")
    ws.merge_cells("C59:E59")
    ws.merge_cells("C60:E60")
    ws.merge_cells("C61:E61")
    ws.merge_cells("I59:K60")
    ws.merge_cells("G61:H61")
    ws.merge_cells("G62:H62")
    ws.merge_cells("I61:J61")

    set_size_cells(ws, min_row=2, max_row=2, min_col=6, max_col=8, size=helv_font_bold)
    set_size_cells(ws, min_row=23, max_row=23, min_col=3, max_col=5, size=helv_font_bold)
    set_size_cells(ws, min_row=24, max_row=24, min_col=3, max_col=5, size=helv_font_bold)
    set_size_cells(ws, min_row=23, max_row=23, min_col=8, max_col=10, size=helv_font_bold)
    set_size_cells(ws, min_row=23, max_row=25, min_col=8, max_col=8, size=helv_font_bold)
    set_size_cells(ws, min_row=24, max_row=25, min_col=8, max_col=8, size=helv_font9_bold)
    set_size_cells(ws, min_row=29, max_row=30, min_col=8, max_col=11, size=helv_font_bold)
    set_size_cells(ws, min_row=33, max_row=33, min_col=3, max_col=11, size=helv_font_bold)
    set_size_cells(ws, min_row=35, max_row=35, min_col=3, max_col=10, size=helv_font_bold)
    set_size_cells(ws, min_row=39, max_row=39, min_col=5, max_col=10, size=helv_font_bold)
    set_size_cells(ws, min_row=46, max_row=47, min_col=3, max_col=9, size=helv_font_bold)
    set_size_cells(ws, min_row=54, max_row=54, min_col=3, max_col=11, size=helv_font_bold)
    set_size_cells(ws, min_row=59, max_row=59, min_col=2, max_col=6, size=helv_font_bold)
    set_size_cells(ws, min_row=59, max_row=59, min_col=6, max_col=6, size=helv_font9_bold)  
    set_size_cells(ws, min_row=8, max_row=8, min_col=6, max_col=6, size=helv_font9)  

    set_alignment_cells(ws, min_row=2, max_row=3, min_col=3, max_col=4, horizontal='right', vertical='center')
    set_alignment_cells(ws, min_row=25, max_row=25, min_col=11, max_col=11, horizontal='left', vertical='center')
    set_alignment_cells(ws, min_row=54, max_row=57, min_col=3, max_col=11, horizontal='left', vertical='center')

    # Aplicar formato específico a celdas que necesitan diferente alineación
    # Celdas con texto largo - alineación izquierda
        
    # Ajustar altura de filas para mejor visualización
    for row in range(1, 70):
        ws.row_dimensions[row].height = 14.25
    
    ws.sheet_view.zoomScale = 70

    ws.add_image(img, "J59")

    # Guardar el archivo
    filename = f"{sAddress}\\{device.fields['so_number'].get()}_purchasing_summary.xlsx"
    filenamePDF = f"{sAddress}\\{device.fields['so_number'].get()}_purchasing_summary.pdf"
    try:
        if 'wb' in locals() and wb is not None:
            wb.save(filename)
    except Exception as e:
        print(f"⚠️ No se pudo guardar el workbook local: {e}")

    # Intenta usar una instancia ya abierta
    try:
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        excel = win32.Dispatch("Excel.Application")

    excel.Visible = True
    excel.DisplayAlerts = False

    # Verificar si el archivo ya está abierto
    workbook = None
    for wb_open in excel.Workbooks:
        try:
            if os.path.abspath(wb_open.FullName).lower() == os.path.abspath(filename).lower():
                workbook = wb_open
                break
        except Exception:
            # Algunos workbooks pueden no tener FullName (ej: libro nuevo)
            pass

    # Si no está abierto, abrirlo
    if workbook is None:
        if not os.path.exists(filename):
            raise FileNotFoundError(f"No se encontró el archivo: {filename}")
        workbook = excel.Workbooks.Open(os.path.abspath(filename))

    if workbook is None:
        raise RuntimeError("No se pudo abrir ni encontrar el archivo de Excel.")

    try:
        ws = workbook.Sheets("Sheet1")
    except Exception:
        # Listar las hojas disponibles si no se encuentra
        sheets = [sh.Name for sh in workbook.Sheets]
        raise ValueError(f"No se encontró la hoja 'Sheet1'. Hojas disponibles: {sheets}")

    ws.PageSetup.PrintArea = "A1:L63"

    # Exportar a PDF
    if os.path.exists(filenamePDF):
        os.remove(filenamePDF)

    ws.ExportAsFixedFormat(
        Type=0,                                # 0 = PDF
        Filename=filenamePDF,
        Quality=0,                             # 0 = Standard, 1 = Minimum
        IncludeDocProperties=True,
        IgnorePrintAreas=False,
        OpenAfterPublish=False
    )

    if not is_folder_open(sAddress):
        os.startfile(sAddress)

    if os.path.exists(filenamePDF):
        os.startfile(filenamePDF)

    return filename

