import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
from datetime import datetime
import win32com.client as win32
import os
from Resources.resources import resource_path
from Definitions.enums import eConnectionType,eArrayType,eLogical,eBIL,eTransformerFamily,eWindingSelection,eTapExtremes,eSequence,eChangerCapacity,eMaterial,eLaminationType,eTypeConductor,eCoating,eDesignSpec,eCoolingType,eSingleCoolingType,eTemperatureRise
from Definitions.functions import address,ADDRESSZ,convertInchCM,is_folder_open

def make_for_ME_page(device = None):

    def set_inside_border(ws,min_row, max_row, min_col, max_col, border, fillE = None):
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border  # apply same border to every cell
                if fillE is not None:
                    cell.fill = fillE

    def set_outer_border(ws, min_row, max_row, min_col, max_col, border, fillE = None):
        """
        Applies a border only around the perimeter of a cell range.

        ws : Worksheet
        min_row : starting row of the range
        max_row : ending row of the range
        min_col : starting column of the range
        max_col : ending column of the range
        border : Border object (e.g., Border(left=Side(...)))
        """
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                # Borde exterior superior
                top = border.top if row == min_row else cell.border.top
                # Borde exterior inferior
                bottom = border.bottom if row == max_row else cell.border.bottom
                # Borde exterior izquierdo
                left = border.left if col == min_col else cell.border.left
                # Borde exterior derecho
                right = border.right if col == max_col else cell.border.right
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
                if fillE is not None:
                    cell.fill = fillE

    def set_size_cells(ws, min_row, max_row, min_col, max_col, size):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = size

    def set_alignment_cells(ws, min_row, max_row, min_col, max_col, horizontal, vertical):
        # Filas del rango
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    device_f = device.fields
    sAddress = ADDRESSZ#address(device_f['so_number'].get())
    so_number = str(device_f['so_number'].get())  # ensure it's a string
    sAddress = os.path.join(sAddress, so_number)
    os.makedirs(sAddress, exist_ok=True)  

    #Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    img_path = resource_path("Resources/logo_reports.png")
    img_pil = PILImage.open(img_path)

    dpi = 96
    width_inch = 1.48
    height_inch = 0.37
    width_px = int(width_inch * dpi)
    height_px = int(height_inch * dpi)

    img_pil = img_pil.resize((width_px, height_px), PILImage.Resampling.LANCZOS)

    img_bytes = BytesIO()
    img_pil.save(img_bytes, format="PNG")
    img_bytes.seek(0)

    img = Image(img_bytes)

    ws.page_margins = PageMargins(
                        left=0.25,
                        right=0.25,
                        top=0.25,
                        bottom=0.25,
                        header=0,
                        footer=0
                    )

    # Definir estilos
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))


    thick_border = Border(left=Side(style='medium'),   # también puedes usar 'thick'
                          right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium')
    )

    dotted_border = Border(
                            left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted')
                        )
    # Fondo de puntitos (pattern fill)
    dot_fill = PatternFill(
                            fill_type="gray125",  # patrón de puntos
                            start_color="000000", # color de los puntos: negro
                            end_color="FFFFFF"    # color de fondo: blanco (o transparente)
                        )
    
    white_fill = PatternFill(
                            fill_type="solid",
                            start_color="FFFFFF",
                            end_color="FFFFFF"
                        )

    gray_fill = PatternFill(
                            start_color="D9D9D9", 
                            end_color="D9D9D9", 
                            fill_type="solid")

    green_fill = PatternFill(
                            start_color="D8E4BC", 
                            end_color="D8E4BC", 
                            fill_type="solid")

    pink_fill = PatternFill(
                            start_color="E6B8B7",
                            end_color="E6B8B7",
                            fill_type="solid")

    yellow_fill = PatternFill(
                            start_color="FFFF00", 
                            end_color="FFFF00", 
                            fill_type="solid")

    helv_font11 = Font(name="Helvetica", size=11)
    helv_font = Font(name="Helvetica", size=10)
    helv_font_bold = Font(name="Helvetica", size=10, bold = True)
    helv_font_bold_under = Font(name="Helvetica", size=10, bold = True,underline="single")
    helv_font9 = Font(name="Helvetica", size=9)
    helv_font9_bold = Font(name="Helvetica", size=9, bold = True)
    helv_font8 = Font(name="Helvetica", size=8)
    helv_font8_bold = Font(name="Helvetica", size=8, bold = True)

    # Ajustar anchos de columna
    column_widths = {
        'A': 29.86, 
        'B': 14.14, 
        'C': 12.57,
        'D': 12.57,
        'E': 12.57,
        'F': 12.57,
        'G': 12.57,
        'H': 12.57
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.print_area = "A1:H53"
    ws.page_setup.scale = 74  # 70%
    # Opcional: orientación y ajuste a ancho/alto
    #ws.page_setup.orientation = 'portrait'  # 'landscape' también
    # Datos de la hoja (A1:L69)
    iNoPhases = device_f['iNoPhases'].get()

    pWindH = device.getHVWind()
    pWindL = device.getLVWind()
    pLCoil = pWindL.coils[0]
    pHCoil = pWindH.coils[0]
    pLCond = pLCoil.conductor[0]
    pHCond = pHCoil.conductor[0]
    coreE = device.core
    mechanicalE = device.mechanical

    pWindH_f = pWindH.fields
    pWindL_f = pWindL.fields
    pHCoil_f = pHCoil.fields
    pLCoil_f = pLCoil.fields
    pHCond_f = pHCond.fields
    pLCond_f = pLCond.fields
    coreE_f = coreE.fields
    mechanicalE_f = mechanicalE.fields

    tMatAluminumL = pLCond_f['tMatAluminum'].get()
    tMatAluminumH = pHCond_f['tMatAluminum'].get()
    sCoilLdMat = pWindL_f['sCoilLdMat'].get()

    dAluminum = (
        (pWindL_f['dWeight'].get() if tMatAluminumL==eMaterial.AL else 0) +
        (pWindH_f['dWeight'].get() if tMatAluminumH==eMaterial.AL else 0) +
        (pWindL_f['dWeight_LBLead'].get()*2.2046 if sCoilLdMat == "Aluminum" else 0) +
        (pWindL_f['dWeight_LBLead'].get()*2.2046 if sCoilLdMat=="Aluminum" else 0)
    )
    if iNoPhases==1:
        dCopper = (
            (pWindL_f['dWeight'].get() if tMatAluminumL==eMaterial.CU else 0) +
            (pWindH_f['dWeight'].get() if tMatAluminumH==eMaterial.CU else 0) +
            (pWindL_f['dWeight_LBLead'].get()*2.2046 if sCoilLdMat == "Copper" else 0) +
            (pWindL_f['dWeight_LBLead'].get()*2.2046 if sCoilLdMat=="Copper" else 0)
        )
    else:
        dCopper = (
            (pWindL_f['dWeight'].get() if tMatAluminumL==eMaterial.CU else 0) +
            (pWindH_f['dWeight'].get() if tMatAluminumH==eMaterial.CU else 0) +
            (pWindL_f['dWeightCoilLeads'].get() if sCoilLdMat == "Copper" else 0) +
            (pWindL_f['dWeightBusBar'].get() if sCoilLdMat=="Copper" else 0)
        )

    dKraft = device_f['dWeightKr0022'].get() + device_f['dWeightKr005'].get() + device_f['dWeightKr010'].get() + device_f['dWeightKr015'].get()
    dCrepe = device_f['dWeightCrepe'].get()
    dPB = coreE_f['dWeightsPB1_16'].get()+ coreE_f['dWeightsPB1_8'].get()
    iKNAN = 1 if device.cooling.IsKNAN() else 0
    sMaterialLeads = "Aluminum" if pWindL_f['tMaterialLeads'].get()==eMaterial.AL else "Copper"
    
    dWindowWidth = coreE_f['dSmallWindowWidth'].get() if iNoPhases==1 else coreE_f['dLargeWindowWidth'].get()
    sCoolingWetSurface = "back wall and 2 side walls oil wetted counted for minipads"
    bHaveTaps = pWindH.have_taps()
    bTapsFCBN = True if bHaveTaps and pWindH.changer.fields['tCapacity'].get() == eChangerCapacity.FCBN else False
    iSteps = pWindH.get_noSteps()-1 if bTapsFCBN else pWindL.getNomPos()
    dHVLineAmp = pWindH_f['pdLineAmp'].get()[iSteps] 
    dHVPhaseAmp = pWindH_f['pdPhaseAmp'].get()[iSteps]
    bCalculoWithRad = mechanicalE_f['bCalculoWithRad'].get()
    dAOR = mechanicalE_f['d_AOI0'].get()[0] if bCalculoWithRad else mechanicalE_f['dAOR'].get()
    iRads = mechanicalE_f['iRadiators'].get() if bCalculoWithRad else 0
    iSections = mechanicalE_f['iObleas'].get() if bCalculoWithRad else 0
    dRadWidth = mechanicalE_f['iRadWidth'].get()/25.4 if bCalculoWithRad else 0
    dEEC = mechanicalE_f['dEEC'].get()/25.4 if bCalculoWithRad else 0
    iQtyFansFA1 = mechanicalE_f['iQtyFansFA1'].get() if bCalculoWithRad else 0
    iQtyFansFA2 = mechanicalE_f['iQtyFansFA2'].get() if bCalculoWithRad else 0
    dTankLength = mechanicalE_f['iTankIntLarge'].get() if bCalculoWithRad else 0
    dTankWidth = mechanicalE_f['iTankIntWidth'].get() if bCalculoWithRad else 0
    dTankHeight = mechanicalE_f['iTankIntHeight'].get() if bCalculoWithRad else 0

    data = [
        ["DONE BY EE:", device_f['issued_by'].get(), None,None, None,None, None,None],
        ["SO NUMBER:", device_f['so_number'].get(), None,None, None,None, None,None],
        ["kVA:", pWindH_f['d_kVA'].get()[0], None,None, None,None, None,None],
        ["DATE:", device_f['original_issue_date'].get(), None,None, None,None, None,None],
        ["INPUT CORE SIZE", None, None,None, None,None, None,None],
        ["A=Window Ht", f"{coreE_f['dWindowHeight'].get():.3f}''", f"{10*convertInchCM(coreE_f['dWindowHeight'].get(),1):.0f} mm",None, None,None, None,None],
        ["B= Ht of core", f"{coreE_f['dOverallCoreHeight'].get():.3f}''", f"{10*convertInchCM(coreE_f['dOverallCoreHeight'].get(),1):.0f} mm",None, None,None, None,None],
        ["C= Window Wdth", f"{dWindowWidth:.3f}''", f"{10*convertInchCM(dWindowWidth,1):.0f} mm",None, None,None, None,None],
        ["D= OA Wdth of core", f"{coreE_f['dCoreLengthWithoutPB'].get():.3f}''", f"{10*convertInchCM(coreE_f['dCoreLengthWithoutPB'].get(),1):.0f} mm",None, None,None, None,None],
        ["E= 1/2 core stack", f"{coreE_f['dStackPerLoop'].get():.3f}''", f"{10*convertInchCM(coreE_f['dStackPerLoop'].get(),1):.0f} mm",None, None,None, None,None],
        ["F= Small Win wdth",  f"{coreE_f['dSmallWindowWidth'].get():.3f}''", f"{10*convertInchCM(coreE_f['dSmallWindowWidth'].get(),1):.0f} mm",None, None,None, None,None],
        ["Strip Width", f"{coreE_f['dSteelWidth'].get():.3f}''", f"{10*convertInchCM(coreE_f['dSteelWidth'].get(),1):.0f} mm",None, None,None, None,None],
        ["L=", f"{coreE_f['dCoreLengthWithPB'].get():.3f}''", f"{10*convertInchCM(coreE_f['dCoreLengthWithPB'].get(),1):.0f} mm",None, None,None, None,None],
        ["Core Weight ", f"{coreE_f['dWeightTwoLoops'].get():.0f} lb", f"{(0.453592*coreE_f['dWeightTwoLoops'].get()):.0f} kg",None, None,None, None,None],
        ["Leg centres", f"{coreE_f['dLegCenters'].get():.3f}''",  f"{10*convertInchCM(coreE_f['dLegCenters'].get(),1):.0f} mm",None, None,None, None,None],
        ["INPUT WINDING SPEC.", None, None,None, None,None, None,None],
        ["Form Width", f"{coreE_f['dFormNose'].get():.3f}''", f"{10*convertInchCM(coreE_f['dFormNose'].get(),1):.0f} mm",None, None,None, None,None],
        ["Form Length", f"{coreE_f['dFormSide'].get():.3f}''", f"{10*convertInchCM(coreE_f['dFormSide'].get(),1):.0f} mm",None, None,None, None,None],
        ["Overcoil", f"{pLCoil_f['dOvercoil'].get():.3f}''", f"{10*convertInchCM(pLCoil_f['dOvercoil'].get(),1):.0f} mm",None, None,None, None,None],
        ["LV Nose", f"{pWindL_f['dNose'].get():.3f}''", f"{10*convertInchCM(pWindL_f['dNose'].get(),1):.0f} mm",None, None,None, None,None],
        ["HV Nose", f"{pWindH_f['dNose'].get():.3f}''", f"{10*convertInchCM(pWindH_f['dNose'].get(),1):.0f} mm",None, None,None, None,None],
        ["Noses", f"{device_f['dNoses'].get():.3f}''", f"{10*convertInchCM(device_f['dNoses'].get(),1):.0f} mm",None, None,None, None,None],
        ["Alum Wt  ", f"{dAluminum:.0f} lb", f"{(0.453592*dAluminum):.0f} kg",None, None,None, None,None],
        ["Copper Wt ", f"{dCopper:.0f} lb", f"{(0.453592*dCopper):.0f} kg",None, None,None, None,None],
        ["INPUT PAPER SUMMARY:", None, None,None, None,None, None,None],
        ["Kraft Weight", f"{dKraft:.0f} lb", f"{(0.453592*dKraft):.0f} kg",None, None,None, None,None],
        ["Crepe Weight", f"{dCrepe:.0f} lb", f"{(0.453592*dCrepe):.0f} kg",None, None,None, None,None],
        ["PB Weight", f"{dPB:.0f} lb", f"{(0.453592*dPB):.0f} kg",None, None,None, None,None],
        ["COOLING", None, None,None, None,None, None,None],
        ["Enter 1 for FR3 fluid, otherwise 0", iKNAN, None,"These Data are subject to change - please enter the cooling data manually", None,None, None,None],
        ["Total Loss Watts", device.losses.fields['dMaxWatts'].get(), None,None, None,None, None,None],
        ["TOR Deg C", f"{device.gradients.fields['dTORFinal'].get():.1f}°", None,None, None,None, None,None],
        ["Average Oil Temp Rise Deg C", f"{dAOR:.1f}°", None,"if you don't have radiators at all. You can get all the necessary", None,None, None,None],
        ["# of Rads", iRads, None,None, None,None, None,None],
        ["Sections", iSections, None,None, None,None, None,None],
        ["Rad Width", f"{dRadWidth:.1f}''", f"{dRadWidth*25.4:.0f} mm","data from COOLING CALC NO RADS sheet.", None,None, None,None],
        ["Rad C to C", f"{dEEC:.0f}''", f"{dEEC*25.4:.0f} mm",None, None,None, None,None],
        ["Fans ONAF1", iQtyFansFA1, None,None, None,None, None,None],
        ["Fans ONAF2", iQtyFansFA2, None,None, None,None, None,None],
        ["BUS BARS", None, None,None, None,None, None,None],
        ["LV Lead Width", f"{pWindL_f['dWidthLeads'].get():.2f}''", f"{10*convertInchCM(pWindL_f['dWidthLeads'].get(),1):.0f} mm",None, None,None, None,None],
        ["LV Lead Thickness", f"{pWindL_f['dThicknessLeads'].get():.2f}''", f"{10*convertInchCM(pWindL_f['dThicknessLeads'].get(),1):.0f} mm",None, None,None, None,None],
        ["Number of LV leads", pWindL_f['iQuantityLeads'].get(), None,None, None,None, None,None],
        ["LV material", sMaterialLeads, None,None, None,None, None,None],
        ["Amperage", None, None,None, None,None, None,None],
        ["LV Line Amperage", f"{pWindL_f['pdLineAmp'].get()[pWindL.getNomPos()]:.2f}", None,None, None,None, None,None],
        ["LV Phase Amperage", f"{pWindL_f['pdPhaseAmp'].get()[pWindL.getNomPos()]:.2f}", None,None, None,None, None,None],
        ["HV Line amperage", f"{dHVLineAmp:.2f}", None,None, None,None, None,None],
        ["HV phase amperage", f"{dHVPhaseAmp:.2f}", None,None, None,None, None,None],
        ["Tank Output Data", None, None,None, None,None, None,None],
        ["Tank Length", f"{dTankLength/25.4:.3f}''", f"{dTankLength:.0f} mm",None, None,None, None,None],
        ["Tank Width", f"{dTankWidth/25.4:.3f}''", f"{dTankWidth:.0f} mm",None, None,None, None,None],
        ["Tank Oil Height", f"{dTankHeight/25.4:.3f}''", f"{dTankHeight:.0f} mm",None, None,None, None,None],
        ["Cooling - Wet surface", None, None,sCoolingWetSurface, None,None, None,None]
    ]
    
    for row in ws.iter_rows(min_row=1, max_row=54, min_col=1, max_col=3):
        for cell in row:
            cell.fill = white_fill

    for row in ws.iter_rows(min_row=1, max_row=54, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill

    for row in ws.iter_rows(min_row=6, max_row=15, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill

    for row in ws.iter_rows(min_row=17, max_row=24, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill
    
    for row in ws.iter_rows(min_row=26, max_row=28, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill

    for row in ws.iter_rows(min_row=30, max_row=39, min_col=2, max_col=8):
        for cell in row:
            cell.fill = pink_fill

    for row in ws.iter_rows(min_row=41, max_row=44, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill

    for row in ws.iter_rows(min_row=46, max_row=49, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill

    for row in ws.iter_rows(min_row=51, max_row=53, min_col=2, max_col=3):
        for cell in row:
            cell.fill = green_fill
    
    for row in ws.iter_rows(min_row=54, max_row=54, min_col=1, max_col=3):
        for cell in row:
            cell.fill = yellow_fill

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=16, max_row=16, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=25, max_row=25, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=29, max_row=29, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=40, max_row=40, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=45, max_row=45, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    for row in ws.iter_rows(min_row=50, max_row=50, min_col=1, max_col=3):
        for cell in row:
            cell.fill = gray_fill

    set_outer_border(ws, min_row=1, max_row=53, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=5, max_row=5, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=16, max_row=16, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=25, max_row=25, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=29, max_row=29, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=40, max_row=40, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=45, max_row=45, min_col=1, max_col=3, border=thick_border) 
    set_outer_border(ws, min_row=50, max_row=50, min_col=1, max_col=3, border=thick_border)
    set_outer_border(ws, min_row=54, max_row=54, min_col=1, max_col=3, border=thick_border)

    # Llenar la hoja con datos y aplicar formato
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = helv_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells("B1:C1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("B3:C3")
    ws.merge_cells("B4:C4")
    ws.merge_cells("A5:C5")
    ws.merge_cells("A16:C16")
    ws.merge_cells("A25:C25")
    ws.merge_cells("A29:C29")
    ws.merge_cells("A40:C40")
    ws.merge_cells("A45:C45")
    ws.merge_cells("A50:C50")
    ws.merge_cells("D30:H32")
    ws.merge_cells("D33:H35")
    ws.merge_cells("D36:H39")

    ws['B4'].number_format = "mm/dd/yyyy"

    set_size_cells(ws, min_row=1, max_row=5, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=6, max_row=54, min_col=2, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=16, max_row=16, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=25, max_row=25, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=29, max_row=29, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=40, max_row=40, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=45, max_row=45, min_col=1, max_col=3, size=helv_font_bold)
    set_size_cells(ws, min_row=50, max_row=50, min_col=1, max_col=3, size=helv_font_bold)

    set_alignment_cells(ws, min_row=1, max_row=54, min_col=1, max_col=1, horizontal='left', vertical='center')
    set_alignment_cells(ws, min_row=54, max_row=54, min_col=4, max_col=4, horizontal='left', vertical='center')
    set_alignment_cells(ws, min_row=30, max_row=39, min_col=5, max_col=8, horizontal='center', vertical='justify')

    for row in range(1, 54):
        ws.row_dimensions[row].height = 14.25
    
    ws.sheet_view.zoomScale = 70

    #ws.add_image(img, "J64")

    # Guardar el archivo
    filename = f"{sAddress}\\{device_f['so_number'].get()}_ME.xlsx"
    if wb is not None:
        wb.save(filename)

    # Intenta usar una instancia ya abierta
    try:
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        excel = win32.Dispatch("Excel.Application")

    excel.Visible = True
    excel.DisplayAlerts = False

    # Verificar si el archivo ya está abierto
    workbook = None
    for wb_open in excel.Workbooks:
        try:
            if os.path.abspath(wb_open.FullName).lower() == os.path.abspath(filename).lower():
                workbook = wb_open
                break
        except Exception:
            # Algunos workbooks pueden no tener FullName (ej: libro nuevo)
            pass

    # Si no está abierto, abrirlo
    if workbook is None:
        if not os.path.exists(filename):
            raise FileNotFoundError(f"No se encontró el archivo: {filename}")
        workbook = excel.Workbooks.Open(os.path.abspath(filename))

    if workbook is None:
        raise RuntimeError("No se pudo abrir ni encontrar el archivo de Excel.")

    try:
        ws = workbook.Sheets("Sheet1")
    except Exception:
        # Listar las hojas disponibles si no se encuentra
        sheets = [sh.Name for sh in workbook.Sheets]
        raise ValueError(f"No se encontró la hoja 'Sheet1'. Hojas disponibles: {sheets}")
    ws.PageSetup.PrintArea = "A1:H54"

    if not is_folder_open(sAddress):
        os.startfile(sAddress)
    # Exportar a PDF
    # ws.ExportAsFixedFormat(0, filenamePDF)  # 0 = PDF

    # if os.path.exists(filenamePDF):
    #     os.startfile(filenamePDF)

    return filename
