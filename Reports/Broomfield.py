import openpyxl
from openpyxl import Workbook
from Definitions.functions import address,ADDRESSZ,PI,is_folder_open
import os

# Variables globales
Pitch = 0.0
TotalTurns = 0.0
TurnsPerLayer = 0.0
NumLayers = 0
SlowTurns = 1.5

InterPaper = 0.0
EndPaper = 0.0
EndLayers = 0

ductLayers = []

wsOut = None
rowOut = 1
iCont = 1
bTaps = False
sLeadsOutTOPCoilHeater = []
sLeadsOutTOPCoil = []

def make_Broomfield(device):
    global Pitch, TotalTurns, TurnsPerLayer, NumLayers
    global InterPaper, EndPaper, EndLayers, bTaps
    global wsOut, rowOut, iCont,ductLayers, sLeadsOutTOPCoilHeater, sLeadsOutTOPCoil
    sAddress = ADDRESSZ#address(device.fields["so_number"].get())
    #documents_folder = Path.home() / "Documents"
    so_number = str(device.fields["so_number"].get())  # ensure it's a string
    sAddress = os.path.join(sAddress, so_number)
    #sAddress = documents_folder / so_number
    os.makedirs(sAddress, exist_ok=True)  
    wb = Workbook()
    ws = wb.active
    #ws.title = "Sheet1"

    filename = f"{sAddress}\\{device.fields['so_number'].get()}_broom.xlsx"

    wb = Workbook()
    #wb = openpyxl.load_workbook(filename)
    ws = wb.active
    wb.remove(ws)

    # Crear hoja de salida o limpiar si existe
    if "Broom" in wb.sheetnames:
        wsOut = wb["Broom"]
        wsOut.delete_rows(1, wsOut.max_row)
    else:
        wsOut = wb.create_sheet("Broom")
    rowOut = 1
    iCont = 1

    pWindH = device.getHVWind()
    pBobH = pWindH.coils[0]
    pCondH = pBobH.conductor[0]

    # Leer datos de Excel (ajustar celdas a tu rango)
    Pitch = pCondH.fields["dTotalWidthTurn"].get()  # HV_Cond_Wdth_Trn
    TotalTurns = pBobH.fields["dTurns"].get()  # HV_Total_trns
    TurnsPerLayer = pBobH.fields["iTurnsLayers"].get()  # HV_Turns_Lyr
    NumLayers = int(pBobH.fields["dTotLayers"].get())  # HV_Tot_Lyrs
    InterPaper = pBobH.fields["dInterPaper"].get()  # Paper
    EndPaper = pBobH.fields["dEndPaper"].get()  # End_Paper_No
    EndLayers = pBobH.fields["iTotNoEndLayers"].get()  # End_Layers
    ductLayers = pBobH.fields["i_Ducts"].get()
    bTaps = pWindH.have_taps()
    if bTaps:
        sLeadsOutTOPCoilHeater = pWindH.changer.fields["sLeadsOutTOPCoilHeater"].get()
        sLeadsOutTOPCoil = pWindH.changer.fields["sLeadsOutTOPCoil"].get()
    else:
        sLeadsOutTOPCoilHeater = []
        sLeadsOutTOPCoil = []

    generate_program()
    #final_margin()
    finish_program()

    wb.save(filename)

    filename_txt = f"{sAddress}\\{so_number}.txt"
    export_txt_msdos(filename_txt)

    if not is_folder_open(sAddress):
        os.startfile(sAddress)

def write_line(txt):
    global rowOut
    wsOut.cell(row=rowOut, column=1, value=txt)
    rowOut += 1

def is_duct_layer(lyr):
    return lyr in ductLayers

def get_paper_thickness(lyr):
    if lyr > (NumLayers - 1 - EndLayers):
        return EndPaper * 1000
    else:
        return InterPaper * 1000

def pause_label(lyr):
    global iCont
    paper = get_paper_thickness(lyr)
    if float(paper).is_integer():
        paper_txt = f"{paper:.0f}"
    else:
        paper_txt = f"{paper:.1f}"
    label = f"{iCont} Pause {paper_txt} mill"
    if is_duct_layer(lyr):
        label += " & Duct"
    write_line(label)
    iCont += 1

def format_turn(v):
    if float(v).is_integer():
        return f"{v:.0f}"
    else:
        return f"{v:.1f}"


def is_even(n):
    return n % 2 == 0

def generate_program():
    global iCont,bTaps
    TurnStop = TurnsPerLayer

    middle = int((NumLayers - 1) / 2)
    write_line(f"{iCont} Pitch {Pitch}")
    iCont += 1
    write_line(f"{iCont} Count Direction Top Going")
    iCont += 1
    write_line(f"{iCont} Home Left")
    iCont += 1
    write_line(f"{iCont} Margin 1.50 Right")
    iCont += 1
    write_line(f"{iCont} Traverse Right")
    iCont += 1
    write_line(f"{iCont} Wind Spiral")
    iCont += 1

    lyr = 1

    while lyr <= NumLayers - 1:
        write_line(f"{iCont} Spindle Foot")
        iCont += 1

        v = TurnStop - SlowTurns
        write_line(f"{iCont} Wait For Turn {format_turn(v)} None")
        iCont += 1
        write_line(f"{iCont} Spindle Slow")
        iCont += 1
        write_line(f"{iCont} Wait For Turn {format_turn(TurnStop)} None")
        iCont += 1
        write_line(f"{iCont} Spindle Stop")
        iCont += 1
        write_line(f"{iCont} Traverse Reverse")
        iCont += 1
        pause_label(lyr)
        TurnStop += TurnsPerLayer

        if lyr == middle and bTaps:

            taps()

            TurnStop += TurnsPerLayer
            lyr += 1

        lyr += 1

    lyr -= 1
    final_margin(lyr)

def taps():

    global iCont, sLeadsOutTOPCoilHeater, sLeadsOutTOPCoil
    
    for i in range(8):

        sLeadsOutTOPCoilE = sLeadsOutTOPCoil[i]

        if sLeadsOutTOPCoilE == "-":
            continue
 
        if sLeadsOutTOPCoilHeater[i]=="B":
            continue

        TurnCirc = float(sLeadsOutTOPCoilE)
        name = sLeadsOutTOPCoilHeater[i] if sLeadsOutTOPCoilHeater[i]!="A" else "A & B"

        write_line(f"{iCont} Spindle Foot")
        iCont += 1

        write_line(f"{iCont} Wait For Turn {format_turn(TurnCirc - SlowTurns)} None")
        iCont += 1

        write_line(f"{iCont} Spindle Slow")
        iCont += 1

        write_line(f"{iCont} Wait For Turn {format_turn(TurnCirc)} None")
        iCont += 1

        write_line(f"{iCont} Spindle Stop")
        iCont += 1

        if name == "A & B":
            write_line(f"{iCont} Traverse Reverse")
            iCont += 1

        pause_taps(name)


# ---------------------------------------------------
# PAUSE TAPS
# ---------------------------------------------------

def pause_taps(tap):

    global iCont

    if tap == "A & B":

        paper = get_paper_thickness(2)
        label = f"{iCont} Pause {format_turn(paper)} mill & Taps {tap}"

    else:

        label = f"{iCont} Pause Tap {tap}"

    write_line(label)

    iCont += 1

def final_margin(lyr):
    global iCont
    TurnStop = NumLayers * TurnsPerLayer
    SpareTurns = TurnStop - TotalTurns
    Move = round((SpareTurns * Pitch) / 2, 3)
    if Move > 0:

        if is_even(lyr):
            write_line(f"{iCont} Margin {Move} Right")
        else:
            write_line(f"{iCont} Margin {Move} Left")

        iCont += 1

def finish_program():
    global iCont
    TurnStop = TotalTurns

    write_line(f"{iCont} Spindle Foot")
    iCont += 1

    write_line(f"{iCont} Wait For Turn {format_turn(TurnStop - SlowTurns)} None")
    iCont += 1

    write_line(f"{iCont} Spindle Slow")
    iCont += 1

    write_line(f"{iCont} Wait For Turn {format_turn(TurnStop)} None")
    iCont += 1

    write_line(f"{iCont} Spindle Stop")
    iCont += 1

    write_line(f"{iCont} Pause GO A ROUND WITH A WINDER")
    iCont += 1

    write_line(f"{iCont} Spindle Foot")
    iCont += 1

    write_line(f"{iCont} Wait For Turn {format_turn(TurnStop + 10)} None")
    iCont += 1

    write_line(f"{iCont} Spindle Stop")
    iCont += 1

    write_line(f"{iCont} End Program")
    iCont += 1

    write_line("˜")

def export_txt_msdos(filename_txt):
    """Genera un archivo TXT MS-DOS (CRLF) sin líneas vacías"""
    with open(filename_txt, "w", newline="\r\n") as f:
        for row in wsOut.iter_rows(min_row=1, max_row=wsOut.max_row, values_only=True):
            cell = row[0]  # solo primera columna
            if cell is not None and str(cell).strip() != "":
                f.write(str(cell) + "\n")